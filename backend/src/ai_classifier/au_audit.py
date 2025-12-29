"""
AU Audit Module - Hardcoded extraction and header validation for Australian customs audit.

This module:
- Skips document classification (dumps all PDFs into the model)
- Extracts header-level data
- Performs header validations (Yes/No/N/A)
- Outputs CSV similar to AU audit spreadsheet format
"""
from __future__ import annotations

import os
import csv
import shutil
import asyncio
import json
from pathlib import Path
from typing import Dict, Any, List, Literal
from pydantic import BaseModel, Field
from pydantic_ai import Agent, BinaryContent
from pydantic_ai.models.google import GoogleModel, GoogleModelSettings
from pydantic_ai.providers.google import GoogleProvider

from .file_manager import get_next_run_id, create_run_directory, create_job_directory
from .util.batch_processor import safe_copy_file

# Maximum number of concurrent job workers
MAX_CONCURRENT_JOBS = 30

# Marker file to indicate a job was successfully audited
AUDIT_COMPLETE_MARKER = ".audit_complete"

# Metadata file to track the run for a grouped folder
RUN_METADATA_FILE = ".au_audit_run.json"


def _save_run_metadata(grouped_folder: Path, run_id: str, run_path: Path, csv_path: Path | None) -> None:
    """Save run metadata to the grouped folder for resume capability."""
    import datetime
    metadata = {
        "run_id": run_id,
        "run_path": str(run_path.resolve()),  # Use absolute path
        "csv_path": str(csv_path.resolve()) if csv_path else None,
        "updated_at": datetime.datetime.now().isoformat()
    }
    metadata_file = grouped_folder / RUN_METADATA_FILE
    metadata_file.write_text(json.dumps(metadata, indent=2))


def _load_run_metadata(grouped_folder: Path) -> Dict[str, Any] | None:
    """Load run metadata from the grouped folder if it exists."""
    metadata_file = grouped_folder / RUN_METADATA_FILE
    if metadata_file.exists():
        try:
            return json.loads(metadata_file.read_text())
        except Exception:
            return None
    return None


def _load_existing_csv_results(csv_path: Path) -> List[Dict[str, str]]:
    """Load existing results from a CSV file."""
    if not csv_path.exists():
        return []
    
    results = []
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(dict(row))
    return results


def clear_audit_markers(grouped_folder: Path, clear_run_metadata: bool = True) -> int:
    """Remove all .audit_complete marker files from job folders."""
    removed = 0
    for job_folder in grouped_folder.iterdir():
        if job_folder.is_dir() and job_folder.name.startswith("job_"):
            marker = job_folder / AUDIT_COMPLETE_MARKER
            if marker.exists():
                marker.unlink()
                removed += 1
    
    if clear_run_metadata:
        metadata_file = grouped_folder / RUN_METADATA_FILE
        if metadata_file.exists():
            metadata_file.unlink()
            print("🧹 Cleared run metadata (will create new run folder)", flush=True)
    
    print(f"🧹 Removed {removed} audit markers from {grouped_folder}", flush=True)
    return removed


def get_audit_status(grouped_folder: Path) -> Dict[str, Any]:
    """Get the audit status of all jobs in a grouped folder."""
    completed = []
    pending = []
    
    for job_folder in sorted(grouped_folder.iterdir()):
        if job_folder.is_dir() and job_folder.name.startswith("job_"):
            job_id = job_folder.name.replace("job_", "")
            marker = job_folder / AUDIT_COMPLETE_MARKER
            if marker.exists():
                completed.append(job_id)
            else:
                pending.append(job_id)
    
    print(f"\n📊 Audit Status for {grouped_folder.name}:", flush=True)
    print(f"   Total jobs: {len(completed) + len(pending)}", flush=True)
    print(f"   ✅ Completed: {len(completed)}", flush=True)
    print(f"   ⏳ Pending: {len(pending)}", flush=True)
    
    return {
        "total": len(completed) + len(pending),
        "completed": len(completed),
        "pending": len(pending),
        "completed_jobs": completed,
        "pending_jobs": pending
    }


# Validation status type
ValidationStatus = Literal["Yes", "No", "N/A"]


class DocumentDetection(BaseModel):
    """Detected document types in the job."""
    has_awb: bool = Field(..., description="True if Air Waybill (AWB/HAWB) document is present")
    has_invoice: bool = Field(..., description="True if Commercial Invoice is present")
    has_entry_print: bool = Field(..., description="True if Entry Print/Customs Declaration is present")
    
    @property
    def is_full_set(self) -> bool:
        """Check if all required documents are present."""
        return self.has_awb and self.has_invoice and self.has_entry_print
    
    @property
    def missing_docs(self) -> List[str]:
        """List of missing document types."""
        missing = []
        if not self.has_awb:
            missing.append("AWB")
        if not self.has_invoice:
            missing.append("Invoice")
        if not self.has_entry_print:
            missing.append("Entry Print")
        return missing


class AUAuditExtraction(BaseModel):
    """Pure extraction fields for AU Audit."""
    audit_month: str = Field(..., description="Entry date - STRICTLY dd/mm/yyyy format (e.g., '01/09/2025', '15/11/2025')")
    dhl_job_number: str = Field("", description="DHL Job Number (starts with B followed by digits) - leave empty if not found")
    waybill_number: str = Field(..., description="HAWB exactly as on entry print (typically 10 digits, no spaces)")
    entry_number: str = Field(..., description="Customs entry number")
    entry_date: str = Field(..., description="Entry date - STRICTLY dd/mm/yyyy format (e.g., '01/09/2025', '15/11/2025')")


class AUAuditHeaderValidation(BaseModel):
    """Header-level validation fields for AU Audit - only fields requiring AI validation."""
    # OC - Owner Code
    oc_correct: ValidationStatus = Field(..., description="OC (Owner Code) Correct? - Selected correct owner code based on documentation")
    oc_reasoning: str = Field("", description="Reasoning for OC validation")
    
    # SC - Supplier Code
    sc_correct: ValidationStatus = Field(..., description="SC (Supplier Code) Correct? - Selected correct supplier code based on documentation")
    sc_reasoning: str = Field("", description="Reasoning for SC validation")
    
    # VALUATION
    valuation_correct: ValidationStatus = Field(..., description="VALUATION Correct? - Correctly declare ITOT per invoices received")
    valuation_reasoning: str = Field("", description="Reasoning for VALUATION validation")
    
    # ORIGIN
    origin_correct: ValidationStatus = Field(..., description="ORIGIN Correct? - Goods line level origin matching shipment paperwork")
    origin_reasoning: str = Field("", description="Reasoning for ORIGIN validation")
    
    # FTA
    fta_correct: ValidationStatus = Field(..., description="FTA Correct? - Claiming preference where noted in customer profile")
    fta_reasoning: str = Field("", description="Reasoning for FTA validation")
    
    # PRS/PRT
    prs_prt_correct: ValidationStatus = Field(..., description="PRS/PRT Correct? - Applying correct preference rule type as per FTA certificates")
    prs_prt_reasoning: str = Field("", description="Reasoning for PRS/PRT validation")
    
    # CURRENCY
    currency_correct: ValidationStatus = Field(..., description="CURRENCY Correct? - Currency correct as per listed currency on invoice")
    currency_reasoning: str = Field("", description="Reasoning for CURRENCY validation")
    
    # INCOTERMS
    incoterms_correct: ValidationStatus = Field(..., description="INCOTERMS Correct? - Declared incoterms as per invoice")
    incoterms_reasoning: str = Field("", description="Reasoning for INCOTERMS validation")
    
    # T & I
    t_i_correct: ValidationStatus = Field(..., description="T & I Correct? - Transport and insurance declared as per notes, documentation, or SOP")
    t_i_reasoning: str = Field("", description="Reasoning for T & I validation")
    
    # OTH/DISC
    oth_disc_correct: ValidationStatus = Field(..., description="OTH/DISC Correct? - Declared discount/additions correctly based on invoice")
    oth_disc_reasoning: str = Field("", description="Reasoning for OTH/DISC validation")


class AUAuditResult(BaseModel):
    """Complete AU Audit Result."""
    status: str = Field("", description="Audit status (filled later)")
    documents: DocumentDetection = Field(..., description="Detected document types")
    extraction: AUAuditExtraction
    header_validation: AUAuditHeaderValidation
    auditor_comments: str = Field("", description="Auditor comments explaining issues")
    auditor: str = Field("DTAL", description="Auditor name")


class AUAuditBatchOutput(BaseModel):
    """Output from the AU audit extraction and validation agent."""
    audit_result: AUAuditResult


_AU_AUDIT_SYSTEM_PROMPT = """
You are an expert customs compliance auditor for DHL Express shipments in Australia.

Your task is to:
1. DETECT which document types are present (AWB, Invoice, Entry Print)
2. EXTRACT key audit metadata from the provided documents
3. VALIDATE specific header-level checks by comparing data across documents

**DOCUMENT DETECTION**:
First, identify which documents are present:
- has_awb: Is there an Air Waybill (AWB/HAWB) document?
- has_invoice: Is there a Commercial Invoice?
- has_entry_print: Is there an Entry Print/Customs Declaration?

**Documents Provided**:
You will receive ALL PDF documents for this job (entry print, commercial invoice, air waybill, etc.). Analyze ALL of them together.

**EXTRACTION RULES**:
1. audit_month: Extract the entry date - format STRICTLY as dd/mm/yyyy (e.g., '01/09/2025', '15/11/2025')
2. dhl_job_number: Extract the DHL Job Number if found (starts with B followed by digits) - leave empty if not found
3. waybill_number: Extract the HAWB exactly as shown on entry print (typically 10 digits, no spaces, e.g., '1234567890')
4. entry_number: Extract the customs entry number
5. entry_date: Extract and format STRICTLY as dd/mm/yyyy (e.g., '01/09/2025', '15/11/2025')

**VALIDATION RULES** (Use "Yes", "No", or "N/A"):
For EACH validation, provide the status and a brief reasoning.

1. **OC Correct? (Owner Code)**: 
   - Selected correct owner code based on documentation
   - Check if importer/owner code matches between entry and invoice/documentation

2. **SC Correct? (Supplier Code)**: 
   - Selected correct supplier code based on documentation

3. **VALUATION Correct?**: 
   - Correctly declare ITOT (Invoice Total) per invoices received
   - Ask customer for proof if AWB value differs from invoice

4. **ORIGIN Correct?**: 
   - Goods line level origin matching shipment paperwork
   - If origin not listed, origin accepted as per supplier headquarters
   - Half points if no more than 1 line error made

5. **FTA Correct?**: 
   - Claiming preference where noted in customer profile
   - If no notes on file, and no prealert received, acceptable to pay duty

6. **PRS/PRT Correct?**: 
   - Applying correct preference rule type e.g. RU/BY additional preference
   - Correct preference as per FTA certificates received

7. **CURRENCY Correct?**: 
   - Currency correct as per listed currency on invoice
   - If multiple currencies, use one which balances ITOT

8. **INCOTERMS Correct?**: 
   - Declared incoterms as per invoice
   - If no incoterms on invoice, FCA acceptable
   - FOB = FCA (treat as equivalent/correct)
   - If using DDP/DAP makes shipment low value, okay to accept FCA terms
   - When incoterms don't match: if (customs value - T&I) < 1000 AUD, mark as correct; otherwise perform normal checks

9. **T & I Correct?**: 
   - Transport and insurance declared as per notes, documentation, or SOP

10. **OTH/DISC Correct?**: 
    - Declared discount/additions correctly based on invoice
    - Mark down if question not asked for unreasonable discounts

**CRITICAL RULES**:
- **NEVER fabricate information** - only extract data visible in documents
- If a field is missing, leave empty or use "N/A"
- Be strict but fair in validation
- "Yes" = Correct, "No" = Error found, "N/A" = Not applicable/cannot determine
- **If only entry print is provided** (no invoice, AWB, or other supporting documents): perform extraction only, set ALL validation fields to "N/A"

Return JSON in the specified format.
"""

_au_audit_agent: Agent | None = None

def _get_au_audit_agent() -> Agent:
    global _au_audit_agent
    if _au_audit_agent is not None:
        return _au_audit_agent
    
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY required")
 
    settings = GoogleModelSettings(
        temperature=0.1, 
        google_thinking_config={'thinking_level': 'high'}
        # google_thinking_config={'thinking_budget': 10000}
    )

    model = GoogleModel('gemini-3-pro-preview', provider=GoogleProvider(api_key=api_key))
    _au_audit_agent = Agent(model=model, instructions=_AU_AUDIT_SYSTEM_PROMPT, output_type=AUAuditBatchOutput, retries=5, model_settings=settings)
    return _au_audit_agent


class TokenUsage:
    def __init__(self, input_tokens: int = 0, output_tokens: int = 0, requests: int = 0):
        self.input_tokens = input_tokens
        self.output_tokens = output_tokens
        self.requests = requests
    
    @property
    def total_tokens(self) -> int:
        return self.input_tokens + self.output_tokens


async def run_au_audit(job_id: str, pdf_files: List[Path], broker_name: str = "", output_job_path: Path | None = None) -> tuple[AUAuditResult, TokenUsage]:
    agent = _get_au_audit_agent()
    
    print(f"\n{'='*80}", flush=True)
    print(f"🇦🇺 AU AUDIT - Job {job_id}", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"Processing {len(pdf_files)} document(s)...", flush=True)
    
    if output_job_path:
        output_job_path.mkdir(parents=True, exist_ok=True)
        for pdf_path in pdf_files:
            if pdf_path.exists():
                dest_path = output_job_path / pdf_path.name
                if safe_copy_file(pdf_path, dest_path):
                    print(f"  📁 Copied: {pdf_path.name} → output", flush=True)
                else:
                    print(f"  ⚠️  Failed to copy: {pdf_path.name}", flush=True)

    prompt = f"""
Analyze ALL the following PDF documents for Job {job_id} and:
1. Extract the required audit metadata
2. Perform header-level validations

{f"Broker name (if not found in documents): {broker_name}" if broker_name else ""}

**IMPORTANT REMINDERS**:
- ONLY extract information that is explicitly visible in the documents
- DO NOT make up, guess, or fabricate any information
- If a field cannot be found, leave it empty

The documents are attached below. Analyze them ALL together to cross-reference information.
"""
    message_parts = [prompt]
    
    for pdf_path in pdf_files:
        if pdf_path.exists():
            pdf_bytes = pdf_path.read_bytes()
            message_parts.append(f"\n**Document: {pdf_path.name}**\n")
            message_parts.append(BinaryContent(data=pdf_bytes, media_type="application/pdf"))
            print(f"  📄 Added: {pdf_path.name} ({len(pdf_bytes):,} bytes)", flush=True)
        else:
            print(f"  ⚠️  File not found: {pdf_path}", flush=True)

    try:
        print(f"\n🔄 Calling Gemini for extraction and validation...", flush=True)
        result = await agent.run(message_parts)
        audit_output: AUAuditBatchOutput = result.output
        usage = result.usage()
        token_usage = TokenUsage(usage.request_tokens or 0, usage.response_tokens or 0, usage.requests or 0)
        
        print(f"\n✅ Audit complete for Job {job_id}", flush=True)
        print(f"   Entry Number: {audit_output.audit_result.extraction.entry_number}", flush=True)
        print(f"   HAWB: {audit_output.audit_result.extraction.waybill_number}", flush=True)
        print(f"   📊 Tokens: input={token_usage.input_tokens:,}, output={token_usage.output_tokens:,}, total={token_usage.total_tokens:,}", flush=True)
        
        # Log validation results with reasoning
        hv = audit_output.audit_result.header_validation
        print(f"\n   Header Validations:", flush=True)
        print(f"   - OC: {hv.oc_correct} ({hv.oc_reasoning[:60] if hv.oc_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - SC: {hv.sc_correct} ({hv.sc_reasoning[:60] if hv.sc_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - VALUATION: {hv.valuation_correct} ({hv.valuation_reasoning[:60] if hv.valuation_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - ORIGIN: {hv.origin_correct} ({hv.origin_reasoning[:60] if hv.origin_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - FTA: {hv.fta_correct} ({hv.fta_reasoning[:60] if hv.fta_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - PRS/PRT: {hv.prs_prt_correct} ({hv.prs_prt_reasoning[:60] if hv.prs_prt_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - CURRENCY: {hv.currency_correct} ({hv.currency_reasoning[:60] if hv.currency_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - INCOTERMS: {hv.incoterms_correct} ({hv.incoterms_reasoning[:60] if hv.incoterms_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - T & I: {hv.t_i_correct} ({hv.t_i_reasoning[:60] if hv.t_i_reasoning else 'No reasoning'}...)", flush=True)
        print(f"   - OTH/DISC: {hv.oth_disc_correct} ({hv.oth_disc_reasoning[:60] if hv.oth_disc_reasoning else 'No reasoning'}...)", flush=True)
        
        if audit_output.audit_result.auditor_comments:
            print(f"\n   💬 Comments: {audit_output.audit_result.auditor_comments}", flush=True)
        
        return audit_output.audit_result, token_usage
    except Exception as e:
        print(f"\n❌ AU Audit failed for Job {job_id}: {e}", flush=True)
        raise


def create_csv_row(audit_result: AUAuditResult) -> Dict[str, str]:
    ext = audit_result.extraction
    hv = audit_result.header_validation
    
    # Clean waybill: remove all spaces
    waybill = ext.waybill_number.replace(" ", "") if ext.waybill_number else ""
    
    # Map Yes -> 1, No -> 0, N/A -> 1 (default to correct if not audited/applicable)
    def val_to_score(v: ValidationStatus) -> str:
        if v == "Yes": return "1"
        if v == "No": return "0"
        return "1"  # Default to 1 for N/A in AU score calculation context

    # AI-validated fields (score is calculated via Excel formula with Weightings sheet)
    row_scores = {
        "OC": val_to_score(hv.oc_correct),
        "SC": val_to_score(hv.sc_correct),
        "VALUATION": val_to_score(hv.valuation_correct),
        "ORIGIN": val_to_score(hv.origin_correct),
        "FTA": val_to_score(hv.fta_correct),
        "PRS/PRT": val_to_score(hv.prs_prt_correct),
        "CURRENCY": val_to_score(hv.currency_correct),
        "INCOTERMS": val_to_score(hv.incoterms_correct),
        "T & I": val_to_score(hv.t_i_correct),
        "OTH/DISC": val_to_score(hv.oth_disc_correct),
        # Always "1" - no AI validation needed
        "CP QUESTIONS": "1",
        "RELATED TRANSACTION": "1",
        "NOTES": "1",
        "AQIS": "1",
        "PERMITS": "1",
        "OTHER": "1",
        # Line-level validations - empty but use 1 for score calculation
        "GST E": "1",
        "CLASS": "1",
        "CONCESSION": "1",
        "UOM/QTY": "1",
    }

    # Score is a formula in Excel, not calculated here
    row = {
        "Month-Year": ext.audit_month,
        "User ID": "",
        "# Of FID": "",
        "BROKER": "",  # Leave empty as requested
        "Entry #": ext.entry_number,
        "WAYBILL #": waybill,
        # AI-validated header fields
        "OC": row_scores["OC"],
        "OC reasoning": hv.oc_reasoning,
        "SC": row_scores["SC"],
        "SC reasoning": hv.sc_reasoning,
        "VALUATION": row_scores["VALUATION"],
        "VALUATION reasoning": hv.valuation_reasoning,
        # Always "1" - no AI validation
        "CP QUESTIONS": "1",
        "RELATED TRANSACTION": "1",
        # AI-validated fields
        "ORIGIN": row_scores["ORIGIN"],
        "ORIGIN reasoning": hv.origin_reasoning,
        "FTA": row_scores["FTA"],
        "FTA reasoning": hv.fta_reasoning,
        "PRS/PRT": row_scores["PRS/PRT"],
        "PRS/PRT reasoning": hv.prs_prt_reasoning,
        # Line-level - leave empty
        "GST E": "",
        # AI-validated fields
        "CURRENCY": row_scores["CURRENCY"],
        "CURRENCY reasoning": hv.currency_reasoning,
        "INCOTERMS": row_scores["INCOTERMS"],
        "INCOTERMS reasoning": hv.incoterms_reasoning,
        "T & I": row_scores["T & I"],
        "T & I reasoning": hv.t_i_reasoning,
        "OTH/DISC": row_scores["OTH/DISC"],
        "OTH/DISC reasoning": hv.oth_disc_reasoning,
        # Line-level - leave empty
        "CLASS": "",
        # Always "1" - no AI validation
        "NOTES": "1",
        # Line-level - leave empty
        "CONCESSION": "",
        "UOM/QTY": "",
        # Always "1" - no AI validation
        "AQIS": "1",
        "PERMITS": "1",
        "OTHER": "1",
        # End columns
        "MAIN ERROR": "",
        "FREE TEXT": audit_result.auditor_comments,
        "GOLD STAR": "",
        "OVERRIDE": "",
        "SUPERVISOR COMMENTS": "",
        "SCORE": "",  # Will be Excel formula
        "ERROR TYPE": "",
        "RISK": "",
        "ENTRY AMENDED": "",
        "DATABASE AMENDED": ""
    }
    return row


def create_csv_file_with_headers(output_path: Path) -> Path:
    # Use dummy to get headers
    dummy = create_csv_row(AUAuditResult(
        documents=DocumentDetection(
            has_awb=True,
            has_invoice=True,
            has_entry_print=True
        ),
        extraction=AUAuditExtraction(
            audit_month="",
            dhl_job_number="",
            waybill_number="",
            entry_number="",
            entry_date=""
        ),
        header_validation=AUAuditHeaderValidation(
            oc_correct="N/A",
            sc_correct="N/A",
            valuation_correct="N/A",
            origin_correct="N/A",
            fta_correct="N/A",
            prs_prt_correct="N/A",
            currency_correct="N/A",
            incoterms_correct="N/A",
            t_i_correct="N/A",
            oth_disc_correct="N/A"
        )
    ))
    fieldnames = list(dummy.keys())
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
    return output_path


def append_csv_row(row: Dict[str, str], output_path: Path) -> None:
    existing_rows = _load_existing_csv_results(output_path)
    wb = row.get("WAYBILL #", "")
    existing_rows = [r for r in existing_rows if r.get("WAYBILL #") != wb]
    existing_rows.append(row)
    
    # Sort by waybill number as requested
    existing_rows.sort(key=lambda x: x.get("WAYBILL #", ""))
    
    fieldnames = list(row.keys())
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(existing_rows)


_xlsx_update_lock = asyncio.Lock()


async def append_xlsx_row(row: Dict[str, str], output_path: Path) -> None:
    async with _xlsx_update_lock:
        try:
            from openpyxl import load_workbook, Workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.comments import Comment
            from openpyxl.utils import get_column_letter
        except ImportError as exc:
            raise ImportError("openpyxl required") from exc
        
        if not output_path.exists():
            wb = Workbook()
            wb.save(output_path)
        
        wb = load_workbook(output_path)
        if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
        
        # Create Weightings sheet FIRST (before formulas reference it)
        if "Weightings" not in wb.sheetnames:
            weightings = wb.create_sheet("Weightings")
            weight_headers = ["OC", "SC", "VALUATION", "CP QUESTIONS", "RELATED TRANSACTION", "ORIGIN", "FTA", "PRS/PRT", 
                              "GST E", "CURRENCY", "INCOTERMS", "T & I", "OTH/DISC", "CLASS", "NOTES", "CONCESSION", "UOM/QTY", "AQIS", "PERMITS", "OTHER"]
            weight_values = [15, 2, 5, 10, 1, 10, 4, 2, 5, 7, 3, 5, 5, 2, 10, 4, 2, 2, 5, 5]
            
            for c, h in enumerate(weight_headers, 1):
                cell = weightings.cell(row=1, column=c, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")
            
            for c, v in enumerate(weight_values, 1):
                weightings.cell(row=2, column=c, value=v)
        
        # Use "AU Audit" as default sheet name since broker is empty
        sheet_name = "AU Audit"
        
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            display_headers = [h for h in list(row.keys()) if not h.endswith("reasoning")]
            for col, h in enumerate(display_headers, 1):
                cell = sheet.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

        display_headers = [h for h in list(row.keys()) if not h.endswith("reasoning")]
        
        # Find SCORE column index
        score_col_idx = display_headers.index("SCORE") + 1 if "SCORE" in display_headers else None
        
        # Check for existing WAYBILL #
        waybill_col = None
        for col, h in enumerate(display_headers, 1):
            if h == "WAYBILL #": waybill_col = col; break
        
        target_row = sheet.max_row + 1
        if waybill_col:
            for r in range(2, sheet.max_row + 1):
                if str(sheet.cell(row=r, column=waybill_col).value) == str(row.get("WAYBILL #")):
                    target_row = r; break

        # Columns that should be numeric (1 or 0)
        numeric_cols = {"OC", "SC", "VALUATION", "CP QUESTIONS", "RELATED TRANSACTION", "ORIGIN", "FTA", 
                        "PRS/PRT", "GST E", "CURRENCY", "INCOTERMS", "T & I", "OTH/DISC", "CLASS", 
                        "NOTES", "CONCESSION", "UOM/QTY", "AQIS", "PERMITS", "OTHER"}
        
        for col, h in enumerate(display_headers, 1):
            if h == "SCORE" and score_col_idx:
                # Use formula for SCORE
                cell = sheet.cell(row=target_row, column=col, value=f"=SUMPRODUCT(G{target_row}:Z{target_row},Weightings!$A$2:$T$2)/SUM(Weightings!$A$2:$T$2)")
            else:
                val = row.get(h, "")
                # Convert numeric strings to integers for score columns
                if h in numeric_cols and val in ("1", "0"):
                    val = int(val)
                cell = sheet.cell(row=target_row, column=col, value=val)
            cell.alignment = Alignment(vertical="center")  # No wrap_text for compact rows
            
            reasoning = row.get(f"{h} reasoning")
            if reasoning:
                cell.comment = Comment(reasoning, "Audit System")
        
        # Set compact row height
        sheet.row_dimensions[target_row].height = 20
        
        # Summary Sheet
        if "Summary" not in wb.sheetnames:
            summary = wb.create_sheet("Summary", 0)
            for col, h in enumerate(["WAYBILL #", "Entry #", "SCORE"], 1):
                cell = summary.cell(row=1, column=col, value=h)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        summary = wb["Summary"]
        s_target = summary.max_row + 1
        for r in range(2, summary.max_row + 1):
            if str(summary.cell(row=r, column=1).value) == str(row.get("WAYBILL #")):
                s_target = r; break
        
        summary.cell(row=s_target, column=1, value=row.get("WAYBILL #"))
        summary.cell(row=s_target, column=2, value=row.get("Entry #"))
        # Reference SCORE formula from AU Audit sheet
        if score_col_idx:
            score_col_letter = get_column_letter(score_col_idx)
            summary.cell(row=s_target, column=3, value=f"='AU Audit'!{score_col_letter}{target_row}")
        summary.row_dimensions[s_target].height = 20

        wb.save(output_path)


def write_audit_xlsx(results: List[Dict[str, str]], output_path: Path) -> Path:
    """Write all results to XLSX, sorted by waybill."""
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
        from openpyxl.utils import get_column_letter
    except ImportError as exc:
        raise ImportError("openpyxl required") from exc
    
    if not results:
        return output_path

    # Sort results by waybill number
    results.sort(key=lambda x: str(x.get("WAYBILL #", "")))
    
    wb = Workbook()
    if "Sheet" in wb.sheetnames: wb.remove(wb["Sheet"])
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    # Weightings Sheet - for SCORE formula reference
    # Column order: OC, SC, VALUATION, CP QUESTIONS, RELATED TRANSACTION, ORIGIN, FTA, PRS/PRT, 
    #               GST E, CURRENCY, INCOTERMS, T & I, OTH/DISC, CLASS, NOTES, CONCESSION, UOM/QTY, AQIS, PERMITS, OTHER
    weightings = wb.create_sheet("Weightings")
    weight_headers = ["OC", "SC", "VALUATION", "CP QUESTIONS", "RELATED TRANSACTION", "ORIGIN", "FTA", "PRS/PRT", 
                      "GST E", "CURRENCY", "INCOTERMS", "T & I", "OTH/DISC", "CLASS", "NOTES", "CONCESSION", "UOM/QTY", "AQIS", "PERMITS", "OTHER"]
    weight_values = [15, 2, 5, 10, 1, 10, 4, 2, 5, 7, 3, 5, 5, 2, 10, 4, 2, 2, 5, 5]
    
    for col, h in enumerate(weight_headers, 1):
        cell = weightings.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    for col, v in enumerate(weight_values, 1):
        weightings.cell(row=2, column=col, value=v)
    
    # Summary Sheet
    summary = wb.create_sheet("Summary", 0)
    summary_headers = ["WAYBILL #", "Entry #", "SCORE"]
    
    for col, h in enumerate(summary_headers, 1):
        cell = summary.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
    
    # Main Data Sheet (all results in one sheet since no broker grouping)
    fieldnames = list(results[0].keys())
    display_headers = [h for h in fieldnames if not h.endswith("reasoning")]
    
    sheet = wb.create_sheet("AU Audit")
    
    # Find column indices for SCORE formula (G:Z maps to columns 7-26 in AU Audit sheet)
    # The validation columns start at column 7 (G) after Month-Year, User ID, # Of FID, BROKER, Entry #, WAYBILL #
    score_col_idx = display_headers.index("SCORE") + 1 if "SCORE" in display_headers else None
    
    for col, h in enumerate(display_headers, 1):
        cell = sheet.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Set header row height
    sheet.row_dimensions[1].height = 30
    
    # Columns that should be numeric (1 or 0)
    numeric_cols = {"OC", "SC", "VALUATION", "CP QUESTIONS", "RELATED TRANSACTION", "ORIGIN", "FTA", 
                    "PRS/PRT", "GST E", "CURRENCY", "INCOTERMS", "T & I", "OTH/DISC", "CLASS", 
                    "NOTES", "CONCESSION", "UOM/QTY", "AQIS", "PERMITS", "OTHER"}
    
    for row_idx, r in enumerate(results, 2):
        for col, h in enumerate(display_headers, 1):
            if h == "SCORE" and score_col_idx:
                # Use formula for SCORE
                cell = sheet.cell(row=row_idx, column=col, value=f"=SUMPRODUCT(G{row_idx}:Z{row_idx},Weightings!$A$2:$T$2)/SUM(Weightings!$A$2:$T$2)")
            else:
                val = r.get(h, "")
                # Convert numeric strings to integers for score columns
                if h in numeric_cols and val in ("1", "0"):
                    val = int(val)
                cell = sheet.cell(row=row_idx, column=col, value=val)
            cell.alignment = Alignment(vertical="center")  # No wrap_text for compact rows
            
            reasoning = r.get(f"{h} reasoning")
            if reasoning:
                cell.comment = Comment(reasoning, "Audit System")
        
        # Set row height to compact value
        sheet.row_dimensions[row_idx].height = 20
    
    # Summary rows with formula reference to AU Audit sheet
    for row_idx, r in enumerate(results, 2):
        summary.cell(row=row_idx, column=1, value=r.get("WAYBILL #"))
        summary.cell(row=row_idx, column=2, value=r.get("Entry #"))
        # Reference SCORE formula from AU Audit sheet
        if score_col_idx:
            score_col_letter = get_column_letter(score_col_idx)
            summary.cell(row=row_idx, column=3, value=f"='AU Audit'!{score_col_letter}{row_idx}")
        summary.row_dimensions[row_idx].height = 20
                
    # Column widths
    for col in range(1, len(display_headers) + 1):
        sheet.column_dimensions[sheet.cell(row=1, column=col).column_letter].width = 12

    wb.save(output_path)
    return output_path


async def process_grouped_jobs_au(grouped_folder: Path, broker_name: str = "", resume_failed_only: bool = False) -> Dict[str, Any]:
    print(f"\n{'='*80}", flush=True)
    print("🇦🇺 AU AUDIT BATCH PROCESSING", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"   Grouped folder: {grouped_folder}", flush=True)
    print(f"   Broker name: {broker_name or '(not specified)'}", flush=True)
    print(f"   Resume failed only: {resume_failed_only}", flush=True)
    
    existing_metadata = _load_run_metadata(grouped_folder)
    metadata_file = grouped_folder / RUN_METADATA_FILE
    
    if existing_metadata:
        run_path_exists = Path(existing_metadata["run_path"]).exists()
        print(f"   Found metadata: {metadata_file}", flush=True)
        print(f"   Previous run: {existing_metadata.get('run_id')}", flush=True)
        print(f"   Run path exists: {run_path_exists}", flush=True)
        
        if run_path_exists:
            run_id = existing_metadata["run_id"]
            run_path = Path(existing_metadata["run_path"])
            print(f"   ✅ Resuming run: {run_id}", flush=True)
        else:
            print(f"   ⚠️  Run path not found, creating new run", flush=True)
            run_id = get_next_run_id()
            run_path = create_run_directory(run_id)
            print(f"   New run ID: {run_id}", flush=True)
    else:
        print(f"   No metadata found at: {metadata_file}", flush=True)
        run_id = get_next_run_id()
        run_path = create_run_directory(run_id)
        print(f"   New run ID: {run_id}", flush=True)

    csv_path = run_path / f"au_audit_combined_{run_id}.csv"
    xlsx_path = run_path / f"au_audit_combined_{run_id}.xlsx"
    incomplete_csv_path = run_path / f"incomplete_docs_{run_id}.csv"
    
    if not csv_path.exists(): create_csv_file_with_headers(csv_path)
    
    # Create incomplete docs CSV with headers
    incomplete_fieldnames = ["Job ID", "WAYBILL #", "Entry #", "Has AWB", "Has Invoice", "Has Entry Print", "Missing Documents"]
    if not incomplete_csv_path.exists():
        with open(incomplete_csv_path, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=incomplete_fieldnames)
            writer.writeheader()

    job_folders = [f for f in grouped_folder.iterdir() if f.is_dir() and f.name.startswith("job_")]
    print(f"   Total job folders found: {len(job_folders)}", flush=True)
    print(f"   Max concurrent jobs: {MAX_CONCURRENT_JOBS}", flush=True)
    print(f"{'='*80}\n", flush=True)
    
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_JOBS)

    async def process_job(folder: Path):
        async with semaphore:
            job_id = folder.name.replace("job_", "")
            if resume_failed_only and (folder / AUDIT_COMPLETE_MARKER).exists():
                print(f"⏭️  Skipping job {job_id} (already complete)", flush=True)
                return {"success": True, "skipped": True, "job_id": job_id}
            
            pdfs = list(folder.glob("*.pdf")) + list(folder.glob("*.PDF"))
            if not pdfs: 
                print(f"⚠️  Job {job_id}: No PDFs found", flush=True)
                return {"success": False, "job_id": job_id, "error": "No PDFs"}
            
            out_job = create_job_directory(run_path, job_id)
            try:
                res, usage = await run_au_audit(job_id, pdfs, broker_name, out_job)
                row = create_csv_row(res)
                
                job_csv = out_job / f"au_audit_{job_id}.csv"
                with open(job_csv, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.DictWriter(f, fieldnames=list(row.keys()))
                    writer.writeheader()
                    writer.writerow(row)
                
                # Incremental append (sorted in append_csv_row)
                append_csv_row(row, csv_path)
                await append_xlsx_row(row, xlsx_path)
                
                # Progressively append incomplete docs
                if not res.documents.is_full_set:
                    waybill = res.extraction.waybill_number.replace(" ", "") if res.extraction.waybill_number else ""
                    incomplete_row = {
                        "Job ID": job_id,
                        "WAYBILL #": waybill,
                        "Entry #": res.extraction.entry_number or "",
                        "Has AWB": "Yes" if res.documents.has_awb else "No",
                        "Has Invoice": "Yes" if res.documents.has_invoice else "No",
                        "Has Entry Print": "Yes" if res.documents.has_entry_print else "No",
                        "Missing Documents": ", ".join(res.documents.missing_docs)
                    }
                    with open(incomplete_csv_path, 'a', newline='', encoding='utf-8') as f:
                        writer = csv.DictWriter(f, fieldnames=incomplete_fieldnames)
                        writer.writerow(incomplete_row)
                
                (folder / AUDIT_COMPLETE_MARKER).write_text(f"Completed: {run_id}")
                return {"success": True, "job_id": job_id, "row": row, "token_usage": usage, "audit_result": res}
            except Exception as e:
                return {"success": False, "job_id": job_id, "error": str(e)}

    tasks = [process_job(f) for f in sorted(job_folders)]
    results = await asyncio.gather(*tasks)
    
    # Collect results and token usage
    successful = 0
    failed = 0
    skipped = 0
    total_input_tokens = 0
    total_output_tokens = 0
    total_requests = 0
    incomplete_doc_jobs: List[Dict[str, Any]] = []  # Jobs missing any required documents
    
    print(f"\n{'='*80}", flush=True)
    print("📊 TOKEN USAGE BY JOB", flush=True)
    print(f"{'='*80}", flush=True)
    
    for result in results:
        job_id = result["job_id"]
        if result.get("skipped"):
            skipped += 1
            # Skipped jobs already printed during processing
        elif result["success"]:
            successful += 1
            # Collect token usage
            if result.get("token_usage"):
                usage = result["token_usage"]
                total_input_tokens += usage.input_tokens
                total_output_tokens += usage.output_tokens
                total_requests += usage.requests
                print(f"   Job {job_id}: input={usage.input_tokens:,}, output={usage.output_tokens:,}, total={usage.total_tokens:,}", flush=True)
            
            # Track ALL incomplete document sets
            audit_result = result.get("audit_result")
            if audit_result and not audit_result.documents.is_full_set:
                waybill = audit_result.extraction.waybill_number.replace(" ", "") if audit_result.extraction.waybill_number else ""
                entry_num = audit_result.extraction.entry_number or ""
                incomplete_doc_jobs.append({
                    "job_id": job_id,
                    "waybill": waybill,
                    "entry_number": entry_num,
                    "has_awb": audit_result.documents.has_awb,
                    "has_invoice": audit_result.documents.has_invoice,
                    "has_entry_print": audit_result.documents.has_entry_print,
                    "missing": audit_result.documents.missing_docs
                })
        else:
            failed += 1
            print(f"   Job {job_id}: FAILED - {result.get('error', 'Unknown error')}", flush=True)
    
    # Final write to ensure sorting and clean formatting
    successful_rows = _load_existing_csv_results(csv_path)
    if successful_rows:
        write_audit_xlsx(successful_rows, xlsx_path)
    
    _save_run_metadata(grouped_folder, run_id, run_path, csv_path)
    
    # Log summary
    total_tokens = total_input_tokens + total_output_tokens
    print(f"\n{'='*80}", flush=True)
    print("🎉 AU AUDIT BATCH COMPLETE", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"   Run ID: {run_id}", flush=True)
    print(f"   Output: {run_path}", flush=True)
    print(f"   Total jobs: {len(job_folders)}", flush=True)
    if skipped > 0:
        print(f"   Skipped (already complete): {skipped}", flush=True)
    print(f"   Processed this run: {successful + failed}", flush=True)
    print(f"   Successful: {successful}", flush=True)
    print(f"   Failed: {failed}", flush=True)
    if failed > 0:
        print(f"\n   💡 TIP: Run again with resume_failed_only=True to retry only failed jobs", flush=True)
    print(f"\n📊 TOTAL TOKEN USAGE:", flush=True)
    print(f"   Input tokens:  {total_input_tokens:,}", flush=True)
    print(f"   Output tokens: {total_output_tokens:,}", flush=True)
    print(f"   Total tokens:  {total_tokens:,}", flush=True)
    print(f"   API requests:  {total_requests}", flush=True)
    
    if csv_path.exists():
        print(f"\n📝 Combined CSV: {csv_path}", flush=True)
    if xlsx_path.exists():
        print(f"📊 Combined XLSX: {xlsx_path}", flush=True)
    
    # Log incomplete document jobs summary - CSV already written progressively
    if incomplete_doc_jobs:
        print(f"\n{'='*80}", flush=True)
        print("📋 INCOMPLETE DOCUMENT SETS (Request from DHL)", flush=True)
        print(f"{'='*80}", flush=True)
        print(f"   Jobs with missing documents: {len(incomplete_doc_jobs)}", flush=True)
        print(f"   {'─'*70}", flush=True)
        for job_info in incomplete_doc_jobs:
            missing_str = ", ".join(job_info["missing"])
            awb_display = job_info["waybill"] if job_info["waybill"] else "(no AWB)"
            print(f"   Job {job_info['job_id']} | AWB: {awb_display} | Missing: {missing_str}", flush=True)
        print(f"   {'─'*70}", flush=True)
        print(f"   📄 Incomplete docs CSV: {incomplete_csv_path}", flush=True)
        print(f"   💡 Send this list to DHL to request missing documents", flush=True)
    
    print(f"{'='*80}", flush=True)
    
    return {
        "run_id": run_id,
        "run_path": str(run_path),
        "total": len(job_folders),
        "successful_jobs": successful,
        "failed_jobs": failed,
        "skipped_jobs": skipped,
        "total_input_tokens": total_input_tokens,
        "total_output_tokens": total_output_tokens,
        "total_tokens": total_tokens,
        "csv_path": str(csv_path),
        "xlsx_path": str(xlsx_path),
        "incomplete_doc_jobs": incomplete_doc_jobs,
        "incomplete_csv_path": str(incomplete_csv_path),
        "results": results
    }

