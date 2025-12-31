"""
NZ Audit Module - Hardcoded extraction and header validation for New Zealand customs audit.

This module:
- Skips document classification (dumps all PDFs into the model)
- Extracts header-level data
- Performs header validations (Yes/No/N/A)
- Outputs CSV similar to Catherine's audit spreadsheet format

Three core concepts:
1. Pure Extraction - Extract audit metadata fields
2. Header Validations - Validate header-level checks (Yes/No/N/A)
3. Line Item Validations - (Not implemented in this module - header only)
"""
from __future__ import annotations

import os
import csv
import shutil
import asyncio
import re
from pathlib import Path
from typing import Dict, Any, List, Literal
from collections import defaultdict
from pydantic import BaseModel, Field
from pydantic_ai import Agent, BinaryContent
from pydantic_ai.models.google import GoogleModel
from pydantic_ai.providers.google import GoogleProvider

from .file_manager import get_next_run_id, create_run_directory, create_job_directory
from .util.batch_processor import safe_copy_file

# Maximum number of concurrent job workers
MAX_CONCURRENT_JOBS = 20

# Marker file to indicate a job was successfully audited
AUDIT_COMPLETE_MARKER = ".audit_complete"

# Metadata file to track the run for a grouped folder
RUN_METADATA_FILE = ".nz_audit_run.json"


def _save_run_metadata(grouped_folder: Path, run_id: str, run_path: Path, csv_path: Path | None) -> None:
    """Save run metadata to the grouped folder for resume capability."""
    import json
    metadata = {
        "run_id": run_id,
        "run_path": str(run_path),
        "csv_path": str(csv_path) if csv_path else None,
        "updated_at": __import__("datetime").datetime.now().isoformat()
    }
    metadata_file = grouped_folder / RUN_METADATA_FILE
    metadata_file.write_text(json.dumps(metadata, indent=2))


def _load_run_metadata(grouped_folder: Path) -> Dict[str, Any] | None:
    """Load run metadata from the grouped folder if it exists."""
    import json
    metadata_file = grouped_folder / RUN_METADATA_FILE
    if metadata_file.exists():
        try:
            return json.loads(metadata_file.read_text())
        except Exception:
            return None
    return None


def _load_existing_csv_results(csv_path: Path) -> List[Dict[str, str]]:
    """Load existing results from a CSV file."""
    if not csv_path.exists():
        return []
    
    results = []
    with open(csv_path, 'r', newline='', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            results.append(dict(row))
    return results


def clear_audit_markers(grouped_folder: Path, clear_run_metadata: bool = True) -> int:
    """
    Remove all .audit_complete marker files from job folders.
    Use this to re-run all jobs from scratch.
    
    Args:
        grouped_folder: Path to the grouped folder containing job_* subfolders
        clear_run_metadata: If True, also clears run metadata so a new run folder is created
        
    Returns:
        Number of markers removed
    """
    removed = 0
    for job_folder in grouped_folder.iterdir():
        if job_folder.is_dir() and job_folder.name.startswith("job_"):
            marker = job_folder / AUDIT_COMPLETE_MARKER
            if marker.exists():
                marker.unlink()
                removed += 1
    
    # Also clear run metadata if requested
    if clear_run_metadata:
        metadata_file = grouped_folder / RUN_METADATA_FILE
        if metadata_file.exists():
            metadata_file.unlink()
            print(f"🧹 Cleared run metadata (will create new run folder)", flush=True)
    
    print(f"🧹 Removed {removed} audit markers from {grouped_folder}", flush=True)
    return removed


def get_audit_status(grouped_folder: Path) -> Dict[str, Any]:
    """
    Get the audit status of all jobs in a grouped folder.
    
    Args:
        grouped_folder: Path to the grouped folder containing job_* subfolders
        
    Returns:
        Dictionary with completed, pending, and total counts
    """
    completed = []
    pending = []
    
    for job_folder in sorted(grouped_folder.iterdir()):
        if job_folder.is_dir() and job_folder.name.startswith("job_"):
            job_id = job_folder.name.replace("job_", "")
            marker = job_folder / AUDIT_COMPLETE_MARKER
            if marker.exists():
                completed.append(job_id)
            else:
                pending.append(job_id)
    
    print(f"\n📊 Audit Status for {grouped_folder.name}:", flush=True)
    print(f"   Total jobs: {len(completed) + len(pending)}", flush=True)
    print(f"   ✅ Completed: {len(completed)}", flush=True)
    print(f"   ⏳ Pending: {len(pending)}", flush=True)
    
    return {
        "total": len(completed) + len(pending),
        "completed": len(completed),
        "pending": len(pending),
        "completed_jobs": completed,
        "pending_jobs": pending
    }


# Validation status type
ValidationStatus = Literal["Yes", "No", "N/A"]


class NZAuditExtraction(BaseModel):
    """
    Pure extraction fields - direct data extraction from documents.
    These are metadata fields extracted from the audit documents.
    """
    # Pure extraction fields
    audit_month: str = Field(..., description="Month the entry was lodged (e.g., 'October', 'November')")
    tl: str = Field("", description="TL field (if applicable, otherwise empty)")
    broker: str = Field(..., description="Broker name who processed the entry")
    dhl_job_number: str = Field(..., description="DHL Job Number (e.g., 'B08902508'), also known as Client Reference Number")
    hawb: str = Field(..., description="House Air Waybill number (e.g., '1104796755')")
    import_export: str = Field(..., description="Whether this is an 'Import' or 'Export' entry")
    entry_number: str = Field(..., description="Customs entry number (e.g., '98942477')")
    entry_date: str = Field(..., description="Entry date in dd/mm/yyyy format")


class NZAuditHeaderValidation(BaseModel):
    """
    Header-level validation fields - Yes/No/N/A for each check.
    These checks compare data across documents (entry print, invoice, AWB).
    Each validation includes a reasoning field explaining the decision.
    """
    # Header validations with reasoning
    client_code_name_correct: ValidationStatus = Field(
        ..., 
        description="Client code/name correct? (IE & EE) - Compare importer/exporter name between entry print and invoice"
    )
    client_code_name_reasoning: str = Field(
        "", 
        description="Brief explanation for client code/name validation (e.g., 'Names match exactly', 'Minor abbreviation difference acceptable')"
    )
    
    supplier_or_cnee_correct: ValidationStatus = Field(
        ...,
        description="For Imports (IE): Supplier code/name correct? For Exports (EE): Consignee name correct? - Compare supplier/consignee between documents"
    )
    supplier_or_cnee_reasoning: str = Field(
        "",
        description="Brief explanation for supplier/consignee validation"
    )
    
    invoice_number_correct: ValidationStatus = Field(
        ...,
        description="Invoice Number Correct? - Check if invoice number on entry print is an EXACT match to the commercial invoice number. Only mark 'Yes' for exact character-for-character matches. Use 'No' if there is any difference (partial match, extra characters, missing characters, different format), or if UNNUM/NONUM is used when actual invoice number exists"
    )
    invoice_number_reasoning: str = Field(
        "",
        description="Brief explanation for invoice number validation - state both invoice numbers found and explain if they are an exact match or not (e.g., 'Entry shows INV-12345, invoice shows INV-12345 - exact match', 'Entry shows 12345, invoice shows INV-12345 - NOT exact match')"
    )
    
    vfd_correct: ValidationStatus = Field(
        ...,
        description="VFD (Value for Duty) Correct? - Compare declared value with invoice total"
    )
    vfd_reasoning: str = Field(
        "",
        description="Brief explanation for VFD validation (e.g., 'VFD matches invoice total', 'VFD differs by $50 due to rounding')"
    )
    
    currency_correct: ValidationStatus = Field(
        ...,
        description="Currency Correct? - Check if currency code matches between entry print and invoice"
    )
    currency_reasoning: str = Field(
        "",
        description="Brief explanation for currency validation"
    )
    
    incoterm_correct: ValidationStatus = Field(
        ...,
        description="Incoterm Correct? - Check if incoterm (FOB, CIF, DDP, EXW, etc.) matches between documents"
    )
    incoterm_reasoning: str = Field(
        "",
        description="Brief explanation for incoterm validation (e.g., 'FOB matches on all documents', 'Entry shows DDP but invoice shows FOB')"
    )
    
    freight_zero_if_inclusive_incoterm: ValidationStatus = Field(
        ...,
        description="If freight inclusive incoterm (CIF, DDP, etc.) and no freight on invoice, is freight zero? Use N/A if not applicable"
    )
    freight_zero_reasoning: str = Field(
        "",
        description="Brief explanation for freight zero validation"
    )
    
    freight_correct: ValidationStatus = Field(
        ...,
        description="Freight correct? (Rate card/ETS). Use N/A if freight is zero or for exports"
    )
    freight_correct_reasoning: str = Field(
        "",
        description="Brief explanation for freight correctness validation"
    )
    
    relationship_indicator_correct: ValidationStatus = Field(
        ...,
        description="Relationship Indicator Correct? (Yes/No) - Check if buyer/seller relationship is correctly declared on the entry. If there is evidence of a potentially related party transaction (e.g., same company names, parent-subsidiary relationship, shared ownership, same director/principals, affiliated companies), the entry MUST declare the relationship. If a related party relationship exists but was NOT declared on the entry, mark as 'No'"
    )
    relationship_indicator_reasoning: str = Field(
        "",
        description="Brief explanation for relationship indicator validation - identify if parties appear related (same company names, affiliated entities, parent-subsidiary, shared ownership) and whether this was properly declared on the entry"
    )

    country_of_export_correct: ValidationStatus = Field(
        ...,
        description="Country of Export correct? - Validate the country of export matches between Invoice/AWB and Entry"
    )
    country_of_export_reasoning: str = Field(
        "",
        description="Brief explanation for country of export validation"
    )

    correct_weight_of_goods: ValidationStatus = Field(
        ...,
        description="Correct weight of goods? - Compare Total Gross/Net Weight on AWB vs Entry"
    )
    correct_weight_reasoning: str = Field(
        "",
        description="Brief explanation for weight validation"
    )

    cgo_correct: ValidationStatus = Field(
        ...,
        description="CGO (for Exports, where applicable) - Customs Goods Order/Clearance status check for exports"
    )
    cgo_reasoning: str = Field(
        "",
        description="Brief explanation for CGO validation"
    )


class NZAuditResult(BaseModel):
    """
    Complete NZ Audit Result combining extraction and validation.
    This represents one row in the output CSV.
    """
    # Status field (can be filled manually later)
    status: str = Field("", description="Audit status (filled later)")
    
    # Pure extraction fields
    extraction: NZAuditExtraction
    
    # Header validation fields
    header_validation: NZAuditHeaderValidation
    
    # Additional audit output fields
    auditor_comments: str = Field("", description="Auditor comments explaining any issues found")
    date_audited: str = Field("", description="Leave empty - will be filled manually by auditor")
    auditor: str = Field("DTAL", description="Auditor name (default: DTAL)")


class NZAuditBatchOutput(BaseModel):
    """Output from the NZ audit extraction and validation."""
    audit_result: NZAuditResult


# System prompt for NZ audit extraction and validation
_NZ_AUDIT_SYSTEM_PROMPT = """
You are an expert customs compliance auditor for DHL Express shipments in New Zealand.

Your task is to:
1. EXTRACT key audit metadata from the provided documents
2. VALIDATE header-level checks by comparing data across documents

**Documents Provided**:
You will receive ALL PDF documents for this job (entry print, commercial invoice, air waybill, and any other supporting documents). Analyze ALL of them together.

**EXTRACTION RULES**:
1. audit_month: Extract the month when the entry was lodged (from entry date)
2. broker: Extract the broker name who processed the entry
3. dhl_job_number: Extract the DHL Job Number (starts with B followed by digits) also known as Client Reference Number
4. hawb: Extract the House Air Waybill number
5. import_export: Determine if this is an "Import" or "Export"
6. entry_number: Extract the customs entry number
7. entry_date: Extract and format as dd/mm/yyyy

**VALIDATION RULES** (Use "Yes", "No", or "N/A"):

For EACH validation, you must provide:
1. The status ("Yes", "No", or "N/A")
2. A brief reasoning explaining your decision (1-2 sentences)

1. **Client code/name correct?**: 
   - Compare the importer/exporter name on entry print with invoice
   - "Yes" if they match (allow minor spelling variations, abbreviations)
   - "No" if clearly different
   - "N/A" if not applicable
   - Reasoning: Explain what you compared and why it matches/doesn't match

2. **Supplier/Consignee correct?**:
   - For IMPORTS: Check if supplier name on entry matches supplier on invoice
   - For EXPORTS: Check if consignee name matches between documents
   - "Yes" if match, "No" if mismatch, "N/A" if not applicable
   - Reasoning: Explain what names you compared and the result

3. **Invoice Number Correct?**:
   - Check if invoice number on entry print is an EXACT character-for-character match to the commercial invoice number
   - "Yes" ONLY if the invoice numbers match exactly (same characters, same format, same case)
   - "No" if there is ANY difference: partial matches, extra characters, missing characters, different format, or if broker used UNNUM/NONUM when an actual invoice number exists
   - "N/A" if not applicable
   - Reasoning: State BOTH invoice numbers found and explicitly confirm if they are an exact match or explain the difference (e.g., 'Entry: INV-12345, Invoice: INV-12345 - exact match' or 'Entry: 12345, Invoice: INV-12345 - NOT exact match, different format')

4. **VFD Correct?**:
   - Compare Value for Duty on entry print with invoice total
   - "Yes" if they match (allow minor rounding)
   - "No" if significant difference
   - "N/A" if not applicable
   - Reasoning: State the VFD value and invoice total, and any difference

5. **Currency Correct?**:
   - Check if currency code matches between entry and invoice
   - "Yes" if match (USD, NZD, etc.)
   - "No" if mismatch
   - Reasoning: State the currencies found in each document

6. **Incoterm Correct?**:
   - Compare incoterm on entry with invoice/AWB
   - "Yes" if correct (FOB, CIF, DDP, EXW, etc.)
   - "No" if mismatch (e.g., invoice says FOB but entry says DDP)
   - Reasoning: State the incoterms found in each document

7. **Freight zero if inclusive incoterm?**:
   - If incoterm is freight-inclusive (CIF, DDP, DDU) AND no freight on invoice
   - Check if freight is declared as zero
   - "Yes" if correctly handled
   - "No" if freight should be zero but isn't
   - "N/A" if not applicable (non-inclusive incoterm or freight shown)
   - Reasoning: Explain the incoterm, whether freight is on invoice, and what was declared

8. **Freight Correct?**:
   - Compare freight charges with rate card/ETS
   - "Yes" if correct
   - "No" if incorrect
   - "N/A" if freight is zero or for exports
   - Reasoning: State the freight amount found and whether it matches expectations

9. **Relationship Indicator Correct?**:
    - Check if buyer/seller relationship is correctly declared on the customs entry
    - Look for evidence of related party transactions: same company names (or variations), parent-subsidiary relationship, shared ownership, same directors/principals, affiliated companies, same address, or any indication of common control
    - "Yes" if either: (a) no relationship exists and none was declared, OR (b) a relationship exists AND it was properly declared on the entry
    - "No" if a potentially related party relationship exists but was NOT declared on the entry (this is the critical error - undeclared related party transactions)
    - "N/A" if not determinable from documents
    - Reasoning: Identify any evidence of related parties (same company names, affiliations, shared ownership, etc.) and confirm whether this was properly declared on the entry. Example: 'Importer XYZ Ltd and Supplier XYZ International appear to be related (similar names/affiliated), but entry shows no relationship declared - should be marked as related party transaction'

11. **Country of Export correct?**:
    - Validate the country of export matches between Invoice/AWB and Entry
    - "Yes" if correct
    - "No" if mismatch
    - Reasoning: State the country of export found in documents

12. **Correct weight of goods?**:
    - Compare Total Gross/Net Weight on AWB vs Entry
    - "Yes" if they match (allow small discrepancies)
    - "No" if significant difference
    - Reasoning: State the weights found

13. **CGO (for Exports)**:
    - For exports, check if CGO (Customs Goods Order) or clearance status is correct/present
    - "Yes" if correct
    - "No" if incorrect/missing
    - "N/A" for imports
    - Reasoning: Explain the finding

**CRITICAL RULES**:
- **NEVER make up or fabricate any information** - only extract data that is explicitly visible in the documents
- If a field cannot be found in the documents, leave it empty or use "N/A" for validations
- If you cannot determine a value with certainty, leave it empty rather than guessing
- **Leave date_audited as an empty string** - this will be filled manually by the auditor
- Be strict but fair in your validation
- When information is missing or unclear, use "N/A" for validation fields
- For exports, freight validation is typically "N/A"
- Include specific reasons for any "No" results in the auditor_comments
- Only report what you can actually see in the documents

Return your result in the exact JSON format specified.
"""


# Cache for agent
_nz_audit_agent: Agent | None = None


def _get_nz_audit_agent() -> Agent:
    """Instantiate (or return cached) Gemini agent for NZ audit."""
    global _nz_audit_agent  # noqa: PLW0603
    
    if _nz_audit_agent is not None:
        return _nz_audit_agent
    
    api_key = os.getenv("GEMINI_API_KEY")
    if not api_key:
        raise EnvironmentError("GEMINI_API_KEY environment variable is required for Gemini agent")

    model = GoogleModel(
        "gemini-3-pro-preview",
        provider=GoogleProvider(api_key=api_key),
    )

    _nz_audit_agent = Agent(
        model=model,
        instructions=_NZ_AUDIT_SYSTEM_PROMPT,
        output_type=NZAuditBatchOutput,
        retries=3,
        model_settings={
            "google_thinking_config": {"thinking_level": "HIGH"},
            "temperature": 0.1
        },
    )
    
    return _nz_audit_agent


class TokenUsage:
    """Token usage tracking for a single job."""
    def __init__(self, input_tokens: int = 0, output_tokens: int = 0, requests: int = 0):
        self.input_tokens = input_tokens
        self.output_tokens = output_tokens
        self.requests = requests
    
    @property
    def total_tokens(self) -> int:
        return self.input_tokens + self.output_tokens
    
    def __repr__(self) -> str:
        return f"TokenUsage(input={self.input_tokens:,}, output={self.output_tokens:,}, total={self.total_tokens:,}, requests={self.requests})"


async def run_nz_audit(
    job_id: str,
    pdf_files: List[Path],
    broker_name: str = "",
    output_job_path: Path | None = None
) -> tuple[NZAuditResult, TokenUsage]:
    """
    Run NZ audit on a job's PDF files.
    
    This function:
    1. Loads all PDF files for the job
    2. Sends them ALL to the model (no classification)
    3. Extracts metadata and performs header validations
    4. Optionally copies files to output folder and saves CSV
    5. Returns structured audit result and token usage
    
    Args:
        job_id: The job ID being audited
        pdf_files: List of paths to PDF files for this job
        broker_name: Optional broker name to pre-fill
        output_job_path: Optional path to output job folder (files will be copied here)
        
    Returns:
        Tuple of (NZAuditResult, TokenUsage)
    """
    agent = _get_nz_audit_agent()
    
    print(f"\n{'='*80}", flush=True)
    print(f"🇳🇿 NZ AUDIT - Job {job_id}", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"Processing {len(pdf_files)} document(s)...", flush=True)
    
    # Copy files to output folder if specified
    if output_job_path:
        output_job_path.mkdir(parents=True, exist_ok=True)
        for pdf_path in pdf_files:
            if pdf_path.exists():
                dest = output_job_path / pdf_path.name
                if safe_copy_file(pdf_path, dest):
                    print(f"  📁 Copied: {pdf_path.name} → output", flush=True)
                else:
                    print(f"  ⚠️  Failed to copy: {pdf_path.name}", flush=True)
    
    # Build message parts - start with prompt
    prompt = f"""
Analyze ALL the following PDF documents for Job {job_id} and:
1. Extract the required audit metadata
2. Perform header-level validations

{f"Broker name (if not found in documents): {broker_name}" if broker_name else ""}

**IMPORTANT REMINDERS**:
- ONLY extract information that is explicitly visible in the documents
- DO NOT make up, guess, or fabricate any information
- If a field cannot be found, leave it empty
- Leave date_audited as an empty string (will be filled manually)

The documents are attached below. Analyze them ALL together to cross-reference information.
"""
    message_parts = [prompt]
    
    # Add all PDF files
    for pdf_path in pdf_files:
        if pdf_path.exists():
            pdf_bytes = pdf_path.read_bytes()
            message_parts.append(f"\n**Document: {pdf_path.name}**\n")
            message_parts.append(BinaryContent(
                data=pdf_bytes,
                media_type="application/pdf"
            ))
            print(f"  📄 Added: {pdf_path.name} ({len(pdf_bytes):,} bytes)", flush=True)
        else:
            print(f"  ⚠️  File not found: {pdf_path}", flush=True)
    
    # Run the audit
    try:
        print(f"\n🔄 Calling Gemini for extraction and validation...", flush=True)
        result = await agent.run(message_parts)
        audit_output: NZAuditBatchOutput = result.output
        
        # Extract token usage
        usage_data = result.usage()
        token_usage = TokenUsage(
            input_tokens=usage_data.request_tokens or 0,
            output_tokens=usage_data.response_tokens or 0,
            requests=usage_data.requests or 0
        )
        
        print(f"\n✅ Audit complete for Job {job_id}", flush=True)
        print(f"   Import/Export: {audit_output.audit_result.extraction.import_export}", flush=True)
        print(f"   Entry Number: {audit_output.audit_result.extraction.entry_number}", flush=True)
        print(f"   HAWB: {audit_output.audit_result.extraction.hawb}", flush=True)
        print(f"   📊 Tokens: input={token_usage.input_tokens:,}, output={token_usage.output_tokens:,}, total={token_usage.total_tokens:,}", flush=True)
        
        # Log validation results with reasoning
        hv = audit_output.audit_result.header_validation
        print(f"\n   Header Validations:", flush=True)
        print(f"   - Client code/name: {hv.client_code_name_correct} ({hv.client_code_name_reasoning[:60] if hv.client_code_name_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Supplier/Cnee: {hv.supplier_or_cnee_correct} ({hv.supplier_or_cnee_reasoning[:60] if hv.supplier_or_cnee_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Invoice Number: {hv.invoice_number_correct} ({hv.invoice_number_reasoning[:60] if hv.invoice_number_reasoning else 'No reasoning'})", flush=True)
        print(f"   - VFD: {hv.vfd_correct} ({hv.vfd_reasoning[:60] if hv.vfd_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Currency: {hv.currency_correct} ({hv.currency_reasoning[:60] if hv.currency_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Incoterm: {hv.incoterm_correct} ({hv.incoterm_reasoning[:60] if hv.incoterm_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Freight Zero: {hv.freight_zero_if_inclusive_incoterm} ({hv.freight_zero_reasoning[:60] if hv.freight_zero_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Freight Correct: {hv.freight_correct} ({hv.freight_correct_reasoning[:60] if hv.freight_correct_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Relationship: {hv.relationship_indicator_correct} ({hv.relationship_indicator_reasoning[:60] if hv.relationship_indicator_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Country of Export: {hv.country_of_export_correct} ({hv.country_of_export_reasoning[:60] if hv.country_of_export_reasoning else 'No reasoning'})", flush=True)
        print(f"   - Weight Correct: {hv.correct_weight_of_goods} ({hv.correct_weight_reasoning[:60] if hv.correct_weight_reasoning else 'No reasoning'})", flush=True)
        print(f"   - CGO Correct: {hv.cgo_correct} ({hv.cgo_reasoning[:60] if hv.cgo_reasoning else 'No reasoning'})", flush=True)
        
        if audit_output.audit_result.auditor_comments:
            print(f"\n   💬 Comments: {audit_output.audit_result.auditor_comments}", flush=True)
        
        return audit_output.audit_result, token_usage
        
    except Exception as e:
        print(f"\n❌ Audit failed for Job {job_id}: {e}", flush=True)
        raise


def create_csv_row(audit_result: NZAuditResult) -> Dict[str, str]:
    """
    Convert NZAuditResult to a CSV row dictionary matching Catherine's format.
    Includes reasoning columns for each validation (for Excel comments).
    
    Returns:
        Dictionary with CSV column names as keys
    """
    ext = audit_result.extraction
    hv = audit_result.header_validation
    
    # Count errors (No results)
    validation_fields = [
        hv.client_code_name_correct,
        hv.supplier_or_cnee_correct,
        hv.invoice_number_correct,
        hv.vfd_correct,
        hv.currency_correct,
        hv.incoterm_correct,
        hv.freight_zero_if_inclusive_incoterm,
        hv.freight_correct,
        hv.relationship_indicator_correct,
        hv.country_of_export_correct,
        hv.correct_weight_of_goods,
        hv.cgo_correct,
    ]
    
    error_count = sum(1 for v in validation_fields if v == "No")
    # Total is count of non-N/A fields
    total_count = sum(1 for v in validation_fields if v != "N/A")
    
    # Find the column indices for the validation range (J to AC)
    # Column J is the 10th column (index 9), AC is column 29 (index 28)
    # We'll use formulas in XLSX, but for CSV we need to calculate
    # The validation columns start at "Client code/name correct?\nIE & EE" which is column J (10)
    # The last validation column before scores is "CGO (for Exports, where applicable)" which is around column AC
    
    return {
        "Status": audit_result.status,
        "Audit Month (month entry lodged)": ext.audit_month,
        "TL": ext.tl,
        "Broker": ext.broker,
        "DHL Job Nmb": ext.dhl_job_number,
        "HAWB": ext.hawb,
        "Import/Export": ext.import_export,
        "Entry Number": ext.entry_number,
        "Entry Date": ext.entry_date,
        "Client code/name correct?\nIE & EE": hv.client_code_name_correct,
        "Client code/name reasoning": hv.client_code_name_reasoning,
        "IE - Supplier code/name correct?\nEE - Cnee name correct?": hv.supplier_or_cnee_correct,
        "Supplier/Cnee reasoning": hv.supplier_or_cnee_reasoning,
        "Invoice Number Correct": hv.invoice_number_correct,
        "Invoice Number reasoning": hv.invoice_number_reasoning,
        "VFD Correct": hv.vfd_correct,
        "VFD reasoning": hv.vfd_reasoning,
        "Currency Correct": hv.currency_correct,
        "Currency reasoning": hv.currency_reasoning,
        "Incoterm Correct": hv.incoterm_correct,
        "Incoterm reasoning": hv.incoterm_reasoning,
        "If freight inclusive incoterm and no freight on invoice, is freight zero?": hv.freight_zero_if_inclusive_incoterm,
        "Freight zero reasoning": hv.freight_zero_reasoning,
        "Freight correct?\nRate card/ETS\nN/A if freight zero\nN/A for exports": hv.freight_correct,
        "Freight correct reasoning": hv.freight_correct_reasoning,
        
        # Line item columns (empty placeholders)
        "Classification Correct": "",
        "Concession": "",
        "Description (actual goods, not description linked with HS Code)": "",
        "Stats Correct": "",
        "Origin Correct": "",
        "Preference": "",
        
        "Country of Export": hv.country_of_export_correct,
        "Country of Export reasoning": hv.country_of_export_reasoning,
        
        "Load Port Air/Sea": "Yes",  # Not AI validated - defaults to Yes
        "Load Port reasoning": "Not AI validated",
        "Relationship Indicator Correct Yes/No?": hv.relationship_indicator_correct,
        "Relationship Indicator reasoning": hv.relationship_indicator_reasoning,
        
        "Correct weight of goods": hv.correct_weight_of_goods,
        "Weight reasoning": hv.correct_weight_reasoning,
        
        "Core value vs. repair value (where applicable)": "N/A",
        
        "CGO (for Exports, where applicable)": hv.cgo_correct,
        "CGO reasoning": hv.cgo_reasoning,
        
        "Date Audited": "",  # Leave blank - will be filled manually by auditor
        "Auditor: Auditor Comments": audit_result.auditor_comments,
        "Auditor": audit_result.auditor,
        # Formulas will be set in XLSX, for CSV we calculate the values
        "Audit Score - Errors": str(error_count),  # Will be formula in XLSX: =COUNTIF($J3:$AC3,"No")
        "Audit Score - Total": str(total_count),  # Will be formula in XLSX: =COUNTIF($J3:$AC3,"<>N/A")
    }


def create_csv_file_with_headers(output_path: Path) -> Path:
    """
    Create a new CSV file with headers only.
    
    Args:
        output_path: Path to output CSV file
        
    Returns:
        Path to the created CSV file
    """
    # Create a dummy row to get fieldnames
    dummy_result = NZAuditResult(
        status="",
        extraction=NZAuditExtraction(
            audit_month="",
            broker="",
            dhl_job_number="",
            hawb="",
            import_export="",
            entry_number="",
            entry_date=""
        ),
        header_validation=NZAuditHeaderValidation(
            client_code_name_correct="N/A",
            supplier_or_cnee_correct="N/A",
            invoice_number_correct="N/A",
            vfd_correct="N/A",
            currency_correct="N/A",
            incoterm_correct="N/A",
            freight_zero_if_inclusive_incoterm="N/A",
            freight_correct="N/A",
            relationship_indicator_correct="N/A",
            country_of_export_correct="N/A",
            correct_weight_of_goods="N/A",
            cgo_correct="N/A"
        )
    )
    dummy_row = create_csv_row(dummy_result)
    fieldnames = list(dummy_row.keys())
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
    
    print(f"📝 Created CSV file with headers: {output_path}", flush=True)
    return output_path


def append_csv_row(row: Dict[str, str], output_path: Path) -> None:
    """
    Append or update a single row in an existing CSV file.
    If a row with the same HAWB exists, it will be replaced (updated).
    Otherwise, the row will be appended.
    
    Args:
        row: Row dictionary from create_csv_row()
        output_path: Path to existing CSV file
    """
    if not output_path.exists():
        raise FileNotFoundError(f"CSV file does not exist: {output_path}")
    
    # Load existing rows
    existing_rows = _load_existing_csv_results(output_path)
    
    # Check if row with same HAWB exists
    hawb = row.get("HAWB", "")
    existing_rows = [r for r in existing_rows if r.get("HAWB") != hawb]
    
    # Add the new/updated row
    existing_rows.append(row)
    
    # Rewrite the entire file
    fieldnames = list(row.keys())
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(existing_rows)


def write_audit_csv(
    results: List[Dict[str, str]],
    output_path: Path
) -> Path:
    """
    Write audit results to CSV file in Catherine's format.
    
    Args:
        results: List of row dictionaries from create_csv_row()
        output_path: Path to output CSV file
        
    Returns:
        Path to the created CSV file
    """
    if not results:
        raise ValueError("No results to write")
    
    # Get column headers from first result
    fieldnames = list(results[0].keys())
    
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.DictWriter(f, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(results)
    
    print(f"\n📝 CSV written: {output_path}", flush=True)
    return output_path


def normalize_broker_name(broker_name: str) -> str:
    """
    Normalize broker name by:
    - Stripping whitespace
    - Converting to title case (first letter uppercase, rest lowercase)
    - Collapsing multiple spaces to single space
    
    Args:
        broker_name: Raw broker name from CSV
        
    Returns:
        Normalized broker name
    """
    if not broker_name:
        return "Unknown"
    
    # Strip whitespace and collapse multiple spaces
    normalized = re.sub(r'\s+', ' ', broker_name.strip())
    # Convert to title case (e.g., "AZHAR ALI" -> "Azhar Ali")
    normalized = normalized.title()
    
    return normalized


# Lock for XLSX file updates (to prevent concurrent write issues)
_xlsx_update_lock = asyncio.Lock()


def create_xlsx_file_with_headers(output_path: Path) -> Path:
    """
    Create a new XLSX file with headers only (empty workbook ready for incremental updates).
    
    Args:
        output_path: Path to output XLSX file
        
    Returns:
        Path to the created XLSX file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")
    
    # Create a dummy row to get fieldnames
    dummy_result = NZAuditResult(
        status="",
        extraction=NZAuditExtraction(
            audit_month="",
            broker="",
            dhl_job_number="",
            hawb="",
            import_export="",
            entry_number="",
            entry_date=""
        ),
        header_validation=NZAuditHeaderValidation(
            client_code_name_correct="N/A",
            supplier_or_cnee_correct="N/A",
            invoice_number_correct="N/A",
            vfd_correct="N/A",
            currency_correct="N/A",
            incoterm_correct="N/A",
            freight_zero_if_inclusive_incoterm="N/A",
            freight_correct="N/A",
            relationship_indicator_correct="N/A",
            country_of_export_correct="N/A",
            correct_weight_of_goods="N/A",
            cgo_correct="N/A"
        )
    )
    dummy_row = create_csv_row(dummy_result)
    
    # Create workbook with summary sheet
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    summary_sheet = wb.create_sheet("Summary", index=0)
    summary_headers = ["DHL Job Nmb", "HAWB", "Broker"]
    
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    summary_sheet.column_dimensions['A'].width = 15
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 25
    
    wb.save(output_path)
    print(f"📊 Created XLSX file with headers: {output_path}", flush=True)
    return output_path


async def append_xlsx_row(row: Dict[str, str], output_path: Path) -> None:
    """
    Append a single row to an existing XLSX file.
    This function loads the existing file, adds the row, and saves it.
    Uses a lock to prevent concurrent write issues.
    
    Args:
        row: Row dictionary from create_csv_row()
        output_path: Path to existing XLSX file
    """
    async with _xlsx_update_lock:
        try:
            from openpyxl import load_workbook
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.comments import Comment
        except ImportError:
            raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")
        
        if not output_path.exists():
            raise FileNotFoundError(f"XLSX file does not exist: {output_path}")
        
        # Load existing workbook
        wb = load_workbook(output_path)
        
        # Get fieldnames from the row
        fieldnames = list(row.keys())
        
        # Map validation columns to their reasoning columns
        validation_reasoning_map = {
            "Client code/name correct?\nIE & EE": "Client code/name reasoning",
            "IE - Supplier code/name correct?\nEE - Cnee name correct?": "Supplier/Cnee reasoning",
            "Invoice Number Correct": "Invoice Number reasoning",
            "VFD Correct": "VFD reasoning",
            "Currency Correct": "Currency reasoning",
            "Incoterm Correct": "Incoterm reasoning",
            "If freight inclusive incoterm and no freight on invoice, is freight zero?": "Freight zero reasoning",
            "Freight correct?\nRate card/ETS\nN/A if freight zero\nN/A for exports": "Freight correct reasoning",
            "Load Port Air/Sea": "Load Port reasoning",
            "Relationship Indicator Correct Yes/No?": "Relationship Indicator reasoning",
            "Country of Export": "Country of Export reasoning",
            "Correct weight of goods": "Weight reasoning",
            "CGO (for Exports, where applicable)": "CGO reasoning",
        }
        
        # Get broker name and normalize it
        broker_raw = row.get("Broker", "").strip()
        broker_normalized = normalize_broker_name(broker_raw)
        
        # Find or create broker sheet
        sheet_name = broker_normalized[:31]
        sheet_name = re.sub(r'[\\/?*\[\]:]', '_', sheet_name)
        
        if sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]
        else:
            sheet = wb.create_sheet(sheet_name)
            # Write headers
            display_headers = [h for h in fieldnames if not h.endswith("reasoning")]
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for col_idx, header in enumerate(display_headers, start=1):
                cell = sheet.cell(row=1, column=col_idx, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            
            # Set column widths
            for col_idx, header in enumerate(display_headers, start=1):
                col_letter = sheet.cell(row=1, column=col_idx).column_letter
                if "Date" in header:
                    sheet.column_dimensions[col_letter].width = 12
                elif "Comments" in header:
                    sheet.column_dimensions[col_letter].width = 50
                elif len(header) > 20:
                    sheet.column_dimensions[col_letter].width = 25
                else:
                    sheet.column_dimensions[col_letter].width = max(len(header) + 2, 12)
        
        # Find next row
        next_row = sheet.max_row + 1
        
        # Find column indices for formulas
        display_headers = [h for h in fieldnames if not h.endswith("reasoning")]
        errors_col_idx = None
        total_col_idx = None
        hawb_col_idx = None
        
        for col_idx, header in enumerate(display_headers, start=1):
            if header == "Audit Score - Errors":
                errors_col_idx = col_idx
            elif header == "Audit Score - Total":
                total_col_idx = col_idx
            elif header == "HAWB":
                hawb_col_idx = col_idx
        
        # Find Excel column letters for validation range
        first_validation_excel_col = None
        col_count = 0
        for header in fieldnames:
            col_count += 1
            if header == "Client code/name correct?\nIE & EE":
                first_validation_excel_col = col_count
                break
        
        def number_to_excel_column(n):
            result = ""
            while n > 0:
                n -= 1
                result = chr(65 + (n % 26)) + result
                n //= 26
            return result
        
        first_col_letter = number_to_excel_column(first_validation_excel_col) if first_validation_excel_col else "J"
        last_col_letter = "AC"
        
        # Check if row with same HAWB already exists in the sheet (BEFORE writing)
        hawb = row.get("HAWB", "")
        existing_row_idx = None
        if hawb_col_idx:
            for row_idx in range(2, sheet.max_row + 1):
                if sheet.cell(row=row_idx, column=hawb_col_idx).value == hawb:
                    existing_row_idx = row_idx
                    break
        
        if existing_row_idx:
            # Update existing row instead of appending
            next_row = existing_row_idx
            # Clear existing row first (values and comments)
            for col_idx in range(1, len(display_headers) + 1):
                cell = sheet.cell(row=next_row, column=col_idx)
                cell.value = None
                cell.comment = None
        else:
            # Append new row
            next_row = sheet.max_row + 1
        
        # Write row data (either updating existing or appending new)
        for col_idx, header in enumerate(display_headers, start=1):
            value = row.get(header, "")
            
            # Set formulas for audit scores
            if col_idx == errors_col_idx:
                formula = f'=COUNTIF(${first_col_letter}{next_row}:${last_col_letter}{next_row},"No")'
                cell = sheet.cell(row=next_row, column=col_idx, value=formula)
            elif col_idx == total_col_idx:
                formula = f'=COUNTIF(${first_col_letter}{next_row}:${last_col_letter}{next_row},"<>N/A")'
                cell = sheet.cell(row=next_row, column=col_idx, value=formula)
            else:
                cell = sheet.cell(row=next_row, column=col_idx, value=value)
            
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            
            # Add comment with reasoning if this is a validation column
            if header in validation_reasoning_map:
                reasoning_header = validation_reasoning_map[header]
                reasoning = row.get(reasoning_header, "").strip()
                if reasoning:
                    comment = Comment(reasoning, "Audit System")
                    cell.comment = comment
        
        # Update summary sheet - check if HAWB exists first
        summary_sheet = wb["Summary"]
        summary_existing_row_idx = None
        for row_idx in range(2, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row_idx, column=2).value == hawb:
                summary_existing_row_idx = row_idx
                break
        
        if summary_existing_row_idx:
            # Update existing summary row
            summary_sheet.cell(row=summary_existing_row_idx, column=1, value=row.get("DHL Job Nmb", ""))
            summary_sheet.cell(row=summary_existing_row_idx, column=2, value=hawb)
            summary_sheet.cell(row=summary_existing_row_idx, column=3, value=broker_normalized)
        else:
            # Append new summary row
            summary_next_row = summary_sheet.max_row + 1
            summary_sheet.cell(row=summary_next_row, column=1, value=row.get("DHL Job Nmb", ""))
            summary_sheet.cell(row=summary_next_row, column=2, value=hawb)
            summary_sheet.cell(row=summary_next_row, column=3, value=broker_normalized)
        
        # Save workbook
        wb.save(output_path)


def write_audit_xlsx(
    results: List[Dict[str, str]],
    output_path: Path
) -> Path:
    """
    Write audit results to XLSX file with:
    - One sheet per broker (normalized)
    - One summary sheet listing all jobs (DHL Job Nmb, HAWB, Broker)
    - Excel cell comments with reasoning for each validation
    
    Args:
        results: List of row dictionaries from create_csv_row()
        output_path: Path to output XLSX file
        
    Returns:
        Path to the created XLSX file
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.comments import Comment
    except ImportError:
        raise ImportError("openpyxl is required for XLSX export. Install with: pip install openpyxl")
    
    if not results:
        raise ValueError("No results to write")
    
    # Get column headers from first result
    fieldnames = list(results[0].keys())
    
    # Map validation columns to their reasoning columns
    validation_reasoning_map = {
        "Client code/name correct?\nIE & EE": "Client code/name reasoning",
        "IE - Supplier code/name correct?\nEE - Cnee name correct?": "Supplier/Cnee reasoning",
        "Invoice Number Correct": "Invoice Number reasoning",
        "VFD Correct": "VFD reasoning",
        "Currency Correct": "Currency reasoning",
        "Incoterm Correct": "Incoterm reasoning",
        "If freight inclusive incoterm and no freight on invoice, is freight zero?": "Freight zero reasoning",
        "Freight correct?\nRate card/ETS\nN/A if freight zero\nN/A for exports": "Freight correct reasoning",
        "Load Port Air/Sea": "Load Port reasoning",
        "Relationship Indicator Correct Yes/No?": "Relationship Indicator reasoning",
        "Country of Export": "Country of Export reasoning",
        "Correct weight of goods": "Weight reasoning",
        "CGO (for Exports, where applicable)": "CGO reasoning",
    }
    
    # Group results by normalized broker name
    broker_groups: Dict[str, List[Dict[str, str]]] = defaultdict(list)
    summary_rows: List[Dict[str, str]] = []
    
    for row in results:
        broker_raw = row.get("Broker", "").strip()
        broker_normalized = normalize_broker_name(broker_raw)
        broker_groups[broker_normalized].append(row)
        
        # Add to summary
        summary_rows.append({
            "DHL Job Nmb": row.get("DHL Job Nmb", ""),
            "HAWB": row.get("HAWB", ""),
            "Broker": broker_normalized
        })
    
    # Create workbook
    wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # Create summary sheet
    summary_sheet = wb.create_sheet("Summary", index=0)
    summary_headers = ["DHL Job Nmb", "HAWB", "Broker"]
    
    # Style summary headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for col_idx, header in enumerate(summary_headers, start=1):
        cell = summary_sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write summary data
    for row_idx, summary_row in enumerate(summary_rows, start=2):
        for col_idx, header in enumerate(summary_headers, start=1):
            summary_sheet.cell(row=row_idx, column=col_idx, value=summary_row.get(header, ""))
    
    # Auto-adjust column widths for summary
    summary_sheet.column_dimensions['A'].width = 15
    summary_sheet.column_dimensions['B'].width = 15
    summary_sheet.column_dimensions['C'].width = 25
    
    # Create one sheet per broker
    for broker_name in sorted(broker_groups.keys()):
        # Sanitize sheet name (Excel has restrictions: max 31 chars, no special chars)
        sheet_name = broker_name[:31]
        # Replace invalid characters
        sheet_name = re.sub(r'[\\/?*\[\]:]', '_', sheet_name)
        
        sheet = wb.create_sheet(sheet_name)
        broker_rows = broker_groups[broker_name]
        
        # Write headers (excluding reasoning columns - they're only for comments)
        display_headers = [h for h in fieldnames if not h.endswith("reasoning")]
        for col_idx, header in enumerate(display_headers, start=1):
            cell = sheet.cell(row=1, column=col_idx, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Find column indices for formulas
        # We need to find the Excel column letters for J (first validation) and AC (last validation)
        # Column J = 10, Column AC = 29 in Excel
        # But we need to account for all columns including reasoning columns to get the correct Excel column letters
        
        errors_col_idx = None
        total_col_idx = None
        
        # Find indices in display_headers (without reasoning columns)
        for col_idx, header in enumerate(display_headers, start=1):
            if header == "Audit Score - Errors":
                errors_col_idx = col_idx
            elif header == "Audit Score - Total":
                total_col_idx = col_idx
        
        # Find the Excel column letters for the validation range
        # Column J (10) is the first validation column: "Client code/name correct?\nIE & EE"
        # Column AC (29) is the last validation column (hardcoded as per user requirement)
        # We need to find column J position, but AC is fixed
        
        first_validation_excel_col = None  # Column J (10)
        
        col_count = 0
        for header in fieldnames:
            col_count += 1
            if header == "Client code/name correct?\nIE & EE":
                # This is column J (10th column)
                first_validation_excel_col = col_count
                break
        
        # Convert column number to Excel column letter
        def number_to_excel_column(n):
            """Convert 1-based column number to Excel column letter (1=A, 10=J, 29=AC)"""
            result = ""
            while n > 0:
                n -= 1
                result = chr(65 + (n % 26)) + result
                n //= 26
            return result
        
        # Hardcode AC as the last column (column 29) as per user requirement
        first_col_letter = number_to_excel_column(first_validation_excel_col) if first_validation_excel_col else "J"
        last_col_letter = "AC"  # Fixed to AC (column 29) as specified
        
        # Write data rows with comments
        for row_idx, row_data in enumerate(broker_rows, start=2):
            for col_idx, header in enumerate(display_headers, start=1):
                value = row_data.get(header, "")
                
                # Set formulas for audit scores if we have the column indices
                if col_idx == errors_col_idx:
                    # Formula: =COUNTIF($J3:$AC3,"No") - using relative row reference
                    formula = f'=COUNTIF(${first_col_letter}{row_idx}:${last_col_letter}{row_idx},"No")'
                    cell = sheet.cell(row=row_idx, column=col_idx, value=formula)
                elif col_idx == total_col_idx:
                    # Formula: =COUNTIF($J3:$AC3,"<>N/A") - using relative row reference
                    formula = f'=COUNTIF(${first_col_letter}{row_idx}:${last_col_letter}{row_idx},"<>N/A")'
                    cell = sheet.cell(row=row_idx, column=col_idx, value=formula)
                else:
                    cell = sheet.cell(row=row_idx, column=col_idx, value=value)
                
                # Wrap text for long cells
                cell.alignment = Alignment(wrap_text=True, vertical="top")
                
                # Add comment with reasoning if this is a validation column
                if header in validation_reasoning_map:
                    reasoning_header = validation_reasoning_map[header]
                    reasoning = row_data.get(reasoning_header, "").strip()
                    if reasoning:
                        # Create comment with reasoning
                        comment = Comment(reasoning, "Audit System")
                        cell.comment = comment
        
        # Auto-adjust column widths (approximate)
        for col_idx, header in enumerate(display_headers, start=1):
            col_letter = sheet.cell(row=1, column=col_idx).column_letter
            # Set reasonable default widths
            if "Date" in header:
                sheet.column_dimensions[col_letter].width = 12
            elif "Comments" in header:
                sheet.column_dimensions[col_letter].width = 50
            elif len(header) > 20:
                sheet.column_dimensions[col_letter].width = 25
            else:
                sheet.column_dimensions[col_letter].width = max(len(header) + 2, 12)
    
    # Save workbook
    wb.save(output_path)
    
    print(f"\n📊 XLSX written: {output_path}", flush=True)
    print(f"   Summary sheet: {len(summary_rows)} jobs", flush=True)
    print(f"   Broker sheets: {len(broker_groups)} brokers", flush=True)
    print(f"   Validation comments added to cells", flush=True)
    for broker_name, rows in sorted(broker_groups.items()):
        print(f"      - {broker_name}: {len(rows)} jobs", flush=True)
    
    return output_path


async def process_grouped_jobs_nz(
    grouped_folder: Path,
    broker_name: str = "",
    resume_failed_only: bool = False
) -> Dict[str, Any]:
    """
    Process all grouped job folders for NZ audit.
    
    This function:
    1. Creates a run directory in output/
    2. Scans the grouped folder for job subfolders
    3. For each job:
       - If resume_failed_only=True, skips jobs with .audit_complete marker
       - Creates job folder in output/
       - Copies PDF files to output job folder
       - Runs the NZ audit
       - Saves individual job CSV in job folder
       - Creates .audit_complete marker in INPUT job folder on success
    4. Outputs a combined CSV with all job results in run folder
    
    Args:
        grouped_folder: Path to the grouped folder (e.g., input/grouped_2025-12-01_120000/)
        broker_name: Optional broker name to pre-fill
        resume_failed_only: If True, skip jobs that have .audit_complete marker (for resuming failed runs)
        
    Returns:
        Dictionary with:
        - run_id: The run identifier
        - run_path: Path to the output run folder
        - total_jobs: Number of jobs processed
        - successful_jobs: Number of successful audits
        - failed_jobs: Number of failed audits
        - skipped_jobs: Number of already-completed jobs skipped
        - csv_path: Path to the combined output CSV
        - job_results: List of individual job results with paths
    """
    print(f"\n{'='*80}", flush=True)
    print(f"🇳🇿 NZ AUDIT BATCH PROCESSING", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"Grouped folder: {grouped_folder}", flush=True)
    if resume_failed_only:
        print(f"🔄 RESUME MODE: Only processing jobs without {AUDIT_COMPLETE_MARKER} marker", flush=True)
    
    if not grouped_folder.exists():
        raise FileNotFoundError(f"Grouped folder not found: {grouped_folder}")
    
    # Check for existing run metadata - ALWAYS reuse if exists
    existing_metadata = _load_run_metadata(grouped_folder)
    existing_results: List[Dict[str, str]] = []
    
    if existing_metadata and Path(existing_metadata["run_path"]).exists():
        # Reuse existing run folder (one run per grouped folder)
        run_id = existing_metadata["run_id"]
        run_path = Path(existing_metadata["run_path"])
        print(f"📂 Using existing run: {run_path}", flush=True)
        
        # Load existing CSV results to merge (only in resume mode to avoid duplicates)
        if resume_failed_only and existing_metadata.get("csv_path"):
            existing_csv = Path(existing_metadata["csv_path"])
            if existing_csv.exists():
                existing_results = _load_existing_csv_results(existing_csv)
                print(f"   Loaded {len(existing_results)} existing results from CSV", flush=True)
    else:
        # Create new run directory (first time processing this grouped folder)
        run_id = get_next_run_id()
        run_path = create_run_directory(run_id)
        print(f"📂 Created new run: {run_path}", flush=True)
    
    print(f"   Run ID: {run_id}", flush=True)
    
    # Create combined CSV and XLSX files upfront (with headers only)
    combined_csv_path = run_path / f"nz_audit_combined_{run_id}.csv"
    combined_xlsx_path = run_path / f"nz_audit_combined_{run_id}.xlsx"
    
    # Create files with headers if they don't exist
    if not combined_csv_path.exists():
        create_csv_file_with_headers(combined_csv_path)
    if not combined_xlsx_path.exists():
        create_xlsx_file_with_headers(combined_xlsx_path)
    
    print(f"📄 Output files ready: {combined_csv_path.name}, {combined_xlsx_path.name}", flush=True)
    
    # Find all job folders
    job_folders = [
        f for f in grouped_folder.iterdir() 
        if f.is_dir() and f.name.startswith("job_")
    ]
    
    if not job_folders:
        raise ValueError(f"No job folders found in {grouped_folder}")
    
    # Count already completed jobs if in resume mode
    if resume_failed_only:
        completed_count = sum(1 for f in job_folders if (f / AUDIT_COMPLETE_MARKER).exists())
        pending_count = len(job_folders) - completed_count
        print(f"Found {len(job_folders)} job folder(s): {completed_count} completed, {pending_count} pending", flush=True)
    else:
        print(f"Found {len(job_folders)} job folder(s)", flush=True)
    
    print(f"🚀 Processing with {MAX_CONCURRENT_JOBS} concurrent workers", flush=True)
    
    # Semaphore to limit concurrent jobs
    semaphore = asyncio.Semaphore(MAX_CONCURRENT_JOBS)
    
    async def process_single_job(job_folder: Path, csv_path: Path, xlsx_path: Path) -> Dict[str, Any]:
        """Process a single job with semaphore-limited concurrency."""
        async with semaphore:
            job_id = job_folder.name.replace("job_", "")
            marker_file = job_folder / AUDIT_COMPLETE_MARKER
            
            # Check if already completed (resume mode)
            if resume_failed_only and marker_file.exists():
                print(f"   ⏭️  Job {job_id} already completed, skipping...", flush=True)
                return {
                    "job_id": job_id,
                    "success": True,
                    "skipped": True,
                    "error": None,
                    "job_folder": None,
                    "csv_path": None,
                    "result": None,
                    "token_usage": None
                }
            
            # Get all PDF files in the job folder
            pdf_files = list(job_folder.glob("*.pdf")) + list(job_folder.glob("*.PDF"))
            
            if not pdf_files:
                print(f"⚠️  No PDF files in {job_folder.name}, skipping...", flush=True)
                return {
                    "job_id": job_id,
                    "success": False,
                    "skipped": False,
                    "error": "No PDF files found",
                    "job_folder": None,
                    "csv_path": None,
                    "result": None,
                    "token_usage": None
                }
            
            # Create job folder in output
            output_job_path = create_job_directory(run_path, job_id)
            
            try:
                # Run audit for this job (files will be copied to output folder)
                audit_result, token_usage = await run_nz_audit(
                    job_id=job_id,
                    pdf_files=pdf_files,
                    broker_name=broker_name,
                    output_job_path=output_job_path
                )
                
                # Convert to CSV row
                row = create_csv_row(audit_result)
                
                # Save individual job CSV
                job_csv_path = output_job_path / f"nz_audit_{job_id}.csv"
                write_audit_csv([row], job_csv_path)
                
                # Append to combined CSV and XLSX files incrementally
                try:
                    append_csv_row(row, csv_path)
                    await append_xlsx_row(row, xlsx_path)
                    print(f"   ✅ Job {job_id} → {output_job_path.name}/ (appended to combined files)", flush=True)
                except Exception as e:
                    print(f"   ⚠️  Job {job_id} completed but failed to append to combined files: {e}", flush=True)
                    # Continue anyway - individual job CSV is saved
                
                # Mark job as complete in INPUT folder
                marker_file.write_text(f"Completed: {run_id}\n")
                
                return {
                    "job_id": job_id,
                    "success": True,
                    "skipped": False,
                    "error": None,
                    "job_folder": str(output_job_path),
                    "csv_path": str(job_csv_path),
                    "result": row,
                    "token_usage": token_usage
                }
                
            except Exception as e:
                print(f"❌ Failed to audit job {job_id}: {e}", flush=True)
                return {
                    "job_id": job_id,
                    "success": False,
                    "skipped": False,
                    "error": str(e),
                    "job_folder": str(output_job_path),
                    "csv_path": None,
                    "result": None,
                    "token_usage": None
                }
    
    # Process all jobs in parallel with limited concurrency
    tasks = [process_single_job(job_folder, combined_csv_path, combined_xlsx_path) for job_folder in sorted(job_folders)]
    job_results = await asyncio.gather(*tasks)
    
    # Collect results and token usage
    all_results: List[Dict[str, str]] = []
    successful = 0
    failed = 0
    skipped = 0
    
    # Track total token usage
    total_input_tokens = 0
    total_output_tokens = 0
    total_requests = 0
    
    print(f"\n{'='*80}", flush=True)
    print(f"📊 TOKEN USAGE BY JOB", flush=True)
    print(f"{'='*80}", flush=True)
    
    for result in job_results:
        job_id = result["job_id"]
        if result.get("skipped"):
            skipped += 1
            # Don't print skipped jobs in token usage section (already printed during processing)
        elif result["success"]:
            successful += 1
            if result["result"]:
                all_results.append(result["result"])
            
            # Collect token usage
            if result["token_usage"]:
                usage = result["token_usage"]
                total_input_tokens += usage.input_tokens
                total_output_tokens += usage.output_tokens
                total_requests += usage.requests
                print(f"   Job {job_id}: input={usage.input_tokens:,}, output={usage.output_tokens:,}, total={usage.total_tokens:,}", flush=True)
        else:
            failed += 1
            print(f"   Job {job_id}: FAILED - {result.get('error', 'Unknown error')}", flush=True)
    
    # Files are already created and updated incrementally, so no need to write them again
    # Just verify they exist
    if combined_csv_path.exists():
        print(f"\n📝 Combined CSV: {combined_csv_path}", flush=True)
    if combined_xlsx_path.exists():
        print(f"📊 Combined XLSX: {combined_xlsx_path}", flush=True)
    
    # Save run metadata for future resume
    _save_run_metadata(grouped_folder, run_id, run_path, combined_csv_path if combined_csv_path.exists() else None)
    
    # Log total token usage
    total_tokens = total_input_tokens + total_output_tokens
    print(f"\n{'='*80}", flush=True)
    print(f"🎉 NZ AUDIT BATCH COMPLETE", flush=True)
    print(f"{'='*80}", flush=True)
    print(f"   Run ID: {run_id}", flush=True)
    print(f"   Output: {run_path}", flush=True)
    print(f"   Total jobs: {len(job_folders)}", flush=True)
    if skipped > 0:
        print(f"   Skipped (already complete): {skipped}", flush=True)
    print(f"   Processed this run: {successful + failed}", flush=True)
    print(f"   Successful: {successful}", flush=True)
    print(f"   Failed: {failed}", flush=True)
    if failed > 0:
        print(f"\n   💡 TIP: Run again with resume_failed_only=True to retry only failed jobs", flush=True)
    print(f"\n📊 TOTAL TOKEN USAGE:", flush=True)
    print(f"   Input tokens:  {total_input_tokens:,}", flush=True)
    print(f"   Output tokens: {total_output_tokens:,}", flush=True)
    print(f"   Total tokens:  {total_tokens:,}", flush=True)
    print(f"   API requests:  {total_requests}", flush=True)
    print(f"{'='*80}", flush=True)
    
    return {
        "run_id": run_id,
        "run_path": str(run_path),
        "total_jobs": len(job_folders),
        "successful_jobs": successful,
        "failed_jobs": failed,
        "skipped_jobs": skipped,
        "csv_path": str(combined_csv_path) if combined_csv_path else None,
        "xlsx_path": str(combined_xlsx_path) if combined_xlsx_path else None,
        "job_results": job_results,
        "results": all_results  # Keep for backward compatibility
    }
