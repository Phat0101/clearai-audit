"""
NZ Audit Summary Module - Generates accuracy and error summary reports from audited Excel files.

This module:
- Reads an Excel file with auditor-completed audit results (one sheet per broker)
- Calculates accuracy for each broker: 1 - (sum of errors / sum of total)
- Counts errors by validation category for each broker
- Outputs a summary Excel file in the requested format
"""
from __future__ import annotations

import re
from pathlib import Path
from typing import Dict, Any, List, Tuple
from collections import defaultdict
from datetime import datetime

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.cell.rich_text import TextBlock, CellRichText
    from openpyxl.cell.text import InlineFont
except ImportError as exc:
    raise ImportError("openpyxl is required for XLSX processing. Install with: pip install openpyxl") from exc


# Mapping of column headers to error category names (display names for the summary)
# These match the validation columns in the NZ audit output
ERROR_CATEGORY_MAPPING = {
    "Client code/name correct?\nIE & EE": "Client code/name incorrect",
    "IE - Supplier code/name correct?\nEE - Cnee name correct?": "Supplier/Consignee incorrect",
    "Invoice Number Correct": "Invoice Number incorrect",
    "VFD Correct": "VFD incorrect",
    "Currency Correct": "Currency:",
    "Incoterm Correct": "Incoterm:",
    "If freight inclusive incoterm and no freight on invoice, is freight zero?": "Freight zero incorrect",
    "Freight correct?\nRate card/ETS\nN/A if freight zero\nN/A for exports": "Freight incorrect",
    "Classification Correct": "Incorrect Tarriff:",
    "Concession": "Invalid concession:",
    "Description (actual goods, not description linked with HS Code)": "Description not as per Comm Inv:",
    "Stats Correct": "Stats:",
    "Origin Correct": "Origin incorrect:",
    "Preference": "Preference incorrect",
    "Country of Export": "Country of Export incorrect",
    "Load Port Air/Sea": "Load Port incorrect",
    "Relationship Indicator Correct Yes/No?": "Relationship Indicator:",
    "Correct weight of goods": "Weight incorrect",
    "CGO (for Exports, where applicable)": "CGO incorrect",
}

# Error categories to display in the summary (in order)
DISPLAY_ERROR_CATEGORIES = [
    "Incorrect parts concession (302913B ) use:",
    "Invalid concession:",
    "Additional tarriff lines rqd:",
    "Incorrect Tarriff:",
    "Description not as per Comm Inv:",
    "Origin incorrect:",
    "Stats:",
    "Incoterm:",
    "Currency:",
    "Relationship Indicator:",
    "Additional:",
]


def normalize_header(header: str) -> str:
    """
    Normalize header text for matching by removing extra whitespace and newlines.
    """
    if not header:
        return ""
    # Replace newlines with space and collapse multiple spaces
    normalized = re.sub(r'\s+', ' ', str(header).strip())
    return normalized


def find_column_index(headers: List[str], target_patterns: List[str]) -> int | None:
    """
    Find column index by matching against multiple possible patterns.
    Returns 0-based index or None if not found.
    """
    for idx, header in enumerate(headers):
        normalized = normalize_header(header)
        for pattern in target_patterns:
            if pattern.lower() in normalized.lower():
                return idx
    return None


def read_broker_sheet(sheet) -> Tuple[List[Dict[str, Any]], List[str]]:
    """
    Read all rows from a broker sheet.
    
    Args:
        sheet: openpyxl worksheet object
        
    Returns:
        Tuple of (list of row dicts, list of headers)
    """
    rows = []
    headers = []
    
    # Read headers from first row
    for cell in sheet[1]:
        headers.append(str(cell.value) if cell.value else "")
    
    # Read data rows
    for row in sheet.iter_rows(min_row=2, values_only=True):
        row_dict = {}
        for col_idx, value in enumerate(row):
            if col_idx < len(headers):
                row_dict[headers[col_idx]] = value
        
        # Skip empty rows
        if any(v for v in row_dict.values() if v):
            rows.append(row_dict)
    
    return rows, headers


def calculate_broker_accuracy(rows: List[Dict[str, Any]], headers: List[str]) -> float:
    """
    Calculate accuracy for a broker: 1 - (sum of errors / sum of total)
    
    Args:
        rows: List of row dictionaries
        headers: List of column headers
        
    Returns:
        Accuracy as a decimal (0.0 to 1.0)
    """
    # Find error and total columns
    errors_col = find_column_index(headers, ["Audit Score - Errors", "Errors"])
    total_col = find_column_index(headers, ["Audit Score - Total", "Total"])
    
    if errors_col is None or total_col is None:
        # Try to calculate from raw values
        return 0.0
    
    total_errors = 0
    total_total = 0
    
    for row in rows:
        errors_val = row.get(headers[errors_col], 0)
        total_val = row.get(headers[total_col], 0)
        
        # Handle formula results and numeric values
        try:
            if isinstance(errors_val, (int, float)):
                total_errors += int(errors_val)
            elif errors_val and str(errors_val).isdigit():
                total_errors += int(errors_val)
        except (ValueError, TypeError):
            pass
        
        try:
            if isinstance(total_val, (int, float)):
                total_total += int(total_val)
            elif total_val and str(total_val).isdigit():
                total_total += int(total_val)
        except (ValueError, TypeError):
            pass
    
    if total_total == 0:
        return 1.0  # No validations = 100% accuracy
    
    return 1.0 - (total_errors / total_total)


def count_errors_by_category(rows: List[Dict[str, Any]], headers: List[str]) -> Tuple[Dict[str, int], Dict[str, int]]:
    """
    Count errors for each validation category, plus additional errors not in predefined categories.
    
    Args:
        rows: List of row dictionaries
        headers: List of column headers
        
    Returns:
        Tuple of (categorized_errors dict, additional_errors dict)
        - categorized_errors: Maps predefined category names to counts
        - additional_errors: Maps column names to counts for uncategorized errors
    """
    error_counts = defaultdict(int)
    additional_errors = defaultdict(int)
    
    # Build set of mapped column headers (normalized)
    mapped_headers_normalized = set()
    for header in ERROR_CATEGORY_MAPPING.keys():
        mapped_headers_normalized.add(normalize_header(header).lower())
    
    # Columns to exclude from additional errors (not validation columns)
    exclude_columns = {
        "status", "audit month", "tl", "broker", "dhl job", "hawb", "import/export",
        "entry number", "entry date", "date audited", "auditor", "comments",
        "audit score", "errors", "total", "reasoning"
    }
    
    for row in rows:
        matched_columns = set()
        
        # Count predefined category errors
        for header, category_name in ERROR_CATEGORY_MAPPING.items():
            normalized_header = normalize_header(header)
            
            for col_header in headers:
                if normalize_header(col_header) == normalized_header or \
                   normalized_header.lower() in normalize_header(col_header).lower():
                    value = row.get(col_header, "")
                    if value and str(value).strip().lower() == "no":
                        error_counts[category_name] += 1
                    matched_columns.add(col_header)
                    break
        
        # Count additional errors (columns with "No" not in predefined categories)
        for col_header in headers:
            if col_header in matched_columns:
                continue
            
            # Skip non-validation columns
            col_lower = normalize_header(col_header).lower()
            if any(excl in col_lower for excl in exclude_columns):
                continue
            
            value = row.get(col_header, "")
            if value and str(value).strip().lower() == "no":
                # Use a cleaned-up column name
                clean_name = normalize_header(col_header)
                # Shorten long names
                if len(clean_name) > 30:
                    clean_name = clean_name[:27] + "..."
                additional_errors[clean_name] += 1
    
    return dict(error_counts), dict(additional_errors)


def process_audit_file(
    input_path: Path,
    output_path: Path,
    month: str
) -> Dict[str, Any]:
    """
    Process an audited Excel file and generate a summary report.
    
    Args:
        input_path: Path to the input Excel file (auditor completed)
        output_path: Path for the output summary Excel file
        month: Month string for the summary (e.g., "Jul-24")
        
    Returns:
        Dictionary with processing results
    """
    if not input_path.exists():
        raise FileNotFoundError(f"Input file not found: {input_path}")
    
    # Load the workbook
    wb = load_workbook(input_path, data_only=True)  # data_only=True to get calculated formula values
    
    broker_results = {}
    all_errors = defaultdict(lambda: defaultdict(int))
    
    # Process each sheet (skip Summary if present)
    for sheet_name in wb.sheetnames:
        if sheet_name.lower() == "summary":
            continue
        
        sheet = wb[sheet_name]
        rows, headers = read_broker_sheet(sheet)
        
        if not rows:
            continue
        
        # Calculate accuracy
        accuracy = calculate_broker_accuracy(rows, headers)
        
        # Count errors by category (returns tuple of categorized and additional)
        error_counts, additional_errors = count_errors_by_category(rows, headers)
        
        broker_results[sheet_name] = {
            "accuracy": accuracy,
            "error_counts": error_counts,
            "additional_errors": additional_errors,
            "total_rows": len(rows)
        }
        
        # Aggregate errors
        for category, count in error_counts.items():
            all_errors[sheet_name][category] = count
    
    # Generate output workbook
    output_wb = Workbook()
    
    # Remove default sheet
    if "Sheet" in output_wb.sheetnames:
        output_wb.remove(output_wb["Sheet"])
    
    # Create summary sheet
    summary_sheet = output_wb.create_sheet("Summary", index=0)
    
    # Styles
    header_fill = PatternFill(start_color="C6D9F0", end_color="C6D9F0", fill_type="solid")
    header_font = Font(bold=True, color="000000")  # Black text on light blue background
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Write header row
    headers = ["Broker", month]
    for col_idx, header in enumerate(headers, start=1):
        cell = summary_sheet.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Write broker accuracy section
    current_row = 2
    brokers_sorted = sorted(broker_results.keys())
    
    for broker_name in brokers_sorted:
        result = broker_results[broker_name]
        accuracy_pct = f"{result['accuracy'] * 100:.0f}%"
        
        cell_broker = summary_sheet.cell(row=current_row, column=1, value=broker_name)
        cell_broker.border = thin_border
        
        cell_accuracy = summary_sheet.cell(row=current_row, column=2, value=accuracy_pct)
        cell_accuracy.alignment = Alignment(horizontal="center")
        cell_accuracy.border = thin_border
        
        current_row += 1
    
    # Write detailed error breakdown for each broker (no spacing row)
    for broker_name in brokers_sorted:
        result = broker_results[broker_name]
        error_counts = result["error_counts"]
        additional_errors = result.get("additional_errors", {})
        
        # Broker header cell (vertically centered)
        broker_header_cell = summary_sheet.cell(row=current_row, column=1, value=broker_name)
        broker_header_cell.font = Font(bold=True)
        broker_header_cell.alignment = Alignment(vertical="center", horizontal="center")
        broker_header_cell.border = thin_border
        
        # Track which error categories are displayed in the predefined list
        displayed_categories = set()
        
        # Define inline fonts for rich text
        red_inline = InlineFont(color="00FF0000")  # Red color (with alpha)
        black_inline = InlineFont(color="00000000")  # Black color
        
        # Build error list and collect additional content
        error_lines = []
        additional_content = ""
        
        for category in DISPLAY_ERROR_CATEGORIES:
            # Special handling for "Additional:" - collect content
            if category == "Additional:":
                # Collect all undisplayed errors from error_counts
                undisplayed_errors = []
                for error_cat, error_count in error_counts.items():
                    if error_count > 0 and error_cat not in displayed_categories:
                        undisplayed_errors.append(f"{error_cat} {error_count}")
                
                # Also add truly additional errors (from unmapped columns)
                for name, cnt in sorted(additional_errors.items()):
                    if cnt > 0:
                        undisplayed_errors.append(f"{name} {cnt}")
                
                additional_content = ", ".join(undisplayed_errors) if undisplayed_errors else ""
                continue
            
            # Map display category to actual category name and find count
            count = 0
            matched_cat = None
            for error_cat, error_count in error_counts.items():
                if category.rstrip(':').lower() in error_cat.lower() or \
                   error_cat.rstrip(':').lower() in category.lower():
                    count = error_count
                    matched_cat = error_cat
                    break
            
            # Track that this category was displayed
            if matched_cat:
                displayed_categories.add(matched_cat)
            
            # Add category with count
            if count > 0:
                error_lines.append((category + " ", str(count), True))  # has count
            else:
                error_lines.append((category, "", False))  # no count
        
        # Build rich text with red counts
        rich_parts = []
        for i, (cat_text, count_text, has_count) in enumerate(error_lines):
            if i > 0:
                rich_parts.append("\n")
            rich_parts.append(TextBlock(black_inline, cat_text))
            if has_count:
                rich_parts.append(TextBlock(red_inline, count_text))
        
        # Add Additional line
        rich_parts.append("\n")
        rich_parts.append(TextBlock(black_inline, "Additional: "))
        if additional_content:
            rich_parts.append(TextBlock(red_inline, additional_content))
        
        # Write error categories in column B using rich text
        error_cell = summary_sheet.cell(row=current_row, column=2)
        error_cell.value = CellRichText(rich_parts)
        error_cell.alignment = Alignment(wrap_text=True, vertical="top")
        # Thicker left border for breakdown cell
        thick_left_border = Border(
            left=Side(style='medium'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        error_cell.border = thick_left_border
        
        # Adjust row height for wrapped content
        # Add extra height (3 rows worth = 45) for the Additional line which can wrap
        base_height = (len(error_lines) + 1) * 15
        additional_height = 30  # 2 rows height for Additional content
        summary_sheet.row_dimensions[current_row].height = base_height + additional_height
        
        current_row += 1  # Move to next row (no spacing between brokers)
    
    # Set column widths
    summary_sheet.column_dimensions['A'].width = 20
    summary_sheet.column_dimensions['B'].width = 50
    
    # Save output
    output_wb.save(output_path)
    
    return {
        "success": True,
        "input_file": str(input_path),
        "output_file": str(output_path),
        "month": month,
        "broker_count": len(broker_results),
        "broker_results": broker_results
    }


def generate_nz_audit_summary(
    input_path: str | Path,
    output_path: str | Path | None = None,
    month: str = ""
) -> Dict[str, Any]:
    """
    Main entry point for generating NZ audit summary.
    
    Args:
        input_path: Path to the auditor-completed Excel file
        output_path: Optional output path (defaults to input_path with _summary suffix)
        month: Month string for display (e.g., "Jul-24")
        
    Returns:
        Dictionary with processing results
    """
    input_path = Path(input_path)
    
    if output_path is None:
        # Generate output path with _summary suffix
        output_path = input_path.parent / f"{input_path.stem}_summary{input_path.suffix}"
    else:
        output_path = Path(output_path)
    
    if not month:
        # Default to current month
        month = datetime.now().strftime("%b-%y")
    
    print("\n" + "=" * 80, flush=True)
    print("🇳🇿 NZ AUDIT SUMMARY GENERATOR", flush=True)
    print("=" * 80, flush=True)
    print(f"Input:  {input_path}", flush=True)
    print(f"Output: {output_path}", flush=True)
    print(f"Month:  {month}", flush=True)
    
    result = process_audit_file(input_path, output_path, month)
    
    print("\n" + "=" * 80, flush=True)
    print("✅ SUMMARY GENERATED", flush=True)
    print("=" * 80, flush=True)
    print(f"Brokers processed: {result['broker_count']}", flush=True)
    
    for broker_name, data in result.get("broker_results", {}).items():
        accuracy_pct = data["accuracy"] * 100
        total_errors = sum(data["error_counts"].values())
        additional_count = sum(data.get("additional_errors", {}).values())
        if additional_count > 0:
            print(f"   - {broker_name}: {accuracy_pct:.1f}% accuracy, {total_errors} categorized + {additional_count} additional errors", flush=True)
        else:
            print(f"   - {broker_name}: {accuracy_pct:.1f}% accuracy, {total_errors} total errors", flush=True)
    
    print(f"\nOutput saved to: {output_path}", flush=True)
    print("=" * 80, flush=True)
    
    return result

